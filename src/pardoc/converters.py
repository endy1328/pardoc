from __future__ import annotations

from dataclasses import dataclass
import base64
from concurrent.futures import ThreadPoolExecutor
import hashlib
import io
from contextlib import contextmanager, redirect_stderr, redirect_stdout
from html import escape
import json
import os
from pathlib import Path
import re
import shutil
import subprocess
import tempfile
from typing import Any


class ConversionError(RuntimeError):
    """Raised when a file cannot be converted."""


CACHE_SCHEMA_VERSION = 10


@dataclass(slots=True)
class ConversionResult:
    source: Path
    text: str
    html: str
    markdown: str = ""
    debug_data: dict[str, Any] | None = None


@dataclass(slots=True)
class ConversionOptions:
    pdf_mode: str = "reconstructed"
    ocr_mode: str = "auto"
    ocr_dpi: int = 200
    ocr_workers: int = 1
    cache_dir: Path | None = None
    page_numbers: set[int] | None = None
    progress_callback: Any | None = None
    analysis_callback: Any | None = None
    debug_overlays: bool = False


@dataclass(slots=True)
class PageAnalysis:
    layout: str
    layout_confidence: float
    text_blocks: int
    image_blocks: int
    drawing_blocks: int
    table_count: int
    text_line_count: int
    text_char_count: int
    dominant_signal: str
    signal_scores: dict[str, float]


def convert_file(path: Path, options: ConversionOptions | None = None) -> ConversionResult:
    options = options or ConversionOptions()
    suffix = path.suffix.lower()

    if suffix == ".docx":
        return _convert_docx(path)
    if suffix == ".doc":
        return _convert_doc(path)
    if suffix == ".pdf":
        return _convert_pdf(path, options)
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return _convert_xlsx(path)
    if suffix == ".xls":
        return _convert_xls(path)
    if suffix in {".txt", ".md", ".csv", ".tsv"}:
        return _convert_plain_text(path)
    if suffix in {".html", ".htm"}:
        return _convert_html(path)

    raise ConversionError(f"Unsupported file type: {suffix or '(no extension)'}")


def render_markdown(result: ConversionResult) -> str:
    if result.markdown.strip():
        return result.markdown.strip()
    markdown = _markdown_from_html(result.html)
    if markdown:
        return markdown
    return result.text.strip()


def render_json(result: ConversionResult) -> str:
    payload = {
        "source": str(result.source),
        "text": result.text,
        "html": result.html,
        "markdown": render_markdown(result),
        "debug": result.debug_data or {},
    }
    return json.dumps(payload, ensure_ascii=False, indent=2)


def _ocr_profile_cache_token(profile: dict[str, Any]) -> str:
    payload = {
        "key": profile.get("key"),
        "psm_candidates": list(profile.get("psm_candidates", [])),
        "variant_policy": profile.get("variant_policy"),
        "word_confidence_floor": profile.get("word_confidence_floor"),
        "threshold": profile.get("threshold"),
    }
    return hashlib.sha256(json.dumps(payload, sort_keys=True).encode("utf-8")).hexdigest()[:16]


def _convert_docx(path: Path) -> ConversionResult:
    try:
        import mammoth
    except ImportError as exc:
        raise ConversionError("DOCX conversion requires the 'mammoth' package.") from exc

    with path.open("rb") as handle:
        html_result = mammoth.convert_to_html(handle)

    with path.open("rb") as handle:
        text_result = mammoth.extract_raw_text(handle)

    html_body = html_result.value.strip()
    warnings = [message.message for message in html_result.messages]
    html = _build_html_document(path.name, html_body or "<p></p>", warnings)
    return ConversionResult(
        source=path,
        text=text_result.value.strip(),
        html=html,
        debug_data={"kind": "docx", "warnings": warnings, "text_length": len(text_result.value.strip())},
    )


def _convert_doc(path: Path) -> ConversionResult:
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if not soffice:
        raise ConversionError(
            "Legacy .doc conversion requires LibreOffice. Install it and ensure 'soffice' is on PATH."
        )

    with tempfile.TemporaryDirectory(prefix="pardoc-doc-") as temp_dir:
        temp_path = Path(temp_dir)
        command = [
            soffice,
            "--headless",
            "--convert-to",
            "docx",
            "--outdir",
            str(temp_path),
            str(path),
        ]
        result = subprocess.run(command, capture_output=True, text=True, check=False)
        if result.returncode != 0:
            detail = (result.stderr or result.stdout).strip()
            raise ConversionError(f"LibreOffice failed to convert .doc: {detail}")

        converted = temp_path / f"{path.stem}.docx"
        if not converted.exists():
            raise ConversionError("LibreOffice reported success but no DOCX file was produced.")

        nested = _convert_docx(converted)
        return ConversionResult(
            source=path,
            text=nested.text,
            html=nested.html,
            markdown=nested.markdown,
            debug_data={"kind": "doc", "converted_via": "libreoffice", "nested": nested.debug_data or {}},
        )


def _convert_pdf(path: Path, options: ConversionOptions) -> ConversionResult:
    try:
        return _convert_pdf_with_pymupdf(path, options)
    except ImportError:
        return _convert_pdf_with_pypdf(path, options)


def _convert_pdf_with_pymupdf(path: Path, options: ConversionOptions) -> ConversionResult:
    try:
        import fitz
    except ImportError as exc:
        raise ImportError from exc

    document = fitz.open(path)
    page_count = len(document)
    html_sections = []
    text_sections = []
    markdown_sections = []
    page_debug: list[dict[str, Any]] = []
    cache_report = _create_cache_report()
    selected_pages = options.page_numbers or set(range(1, page_count + 1))
    total_selected = len(selected_pages)
    processed = 0
    ocr_cache = _precompute_document_ocr(document, path, selected_pages, options, cache_report)
    try:
        for index, page in enumerate(document, start=1):
            if index not in selected_pages:
                continue
            cached_payload = ocr_cache.get(index)
            text, html, markdown, status, debug = _extract_pymupdf_page(
                page,
                index,
                options,
                path,
                document,
                cache_report,
                cached_payload,
            )
            text_sections.append(f"[Page {index}]\n{text}".strip())
            html_sections.append(html)
            markdown_sections.append(markdown)
            page_debug.append(debug)
            processed += 1
            if options.progress_callback:
                options.progress_callback(index, page_count, processed, total_selected, status)
            if options.analysis_callback:
                options.analysis_callback(debug)
    finally:
        document.close()

    html = _build_html_document(path.name, "\n".join(html_sections) or "<p></p>")
    text = "\n\n".join(section for section in text_sections if section).strip()
    markdown = "\n\n".join(section for section in markdown_sections if section).strip()
    return ConversionResult(
        source=path,
        text=text,
        html=html,
        markdown=markdown,
        debug_data={
            "kind": "pdf",
            "engine": "pymupdf",
            "page_count": page_count,
            "selected_pages": sorted(selected_pages),
            "cache": _finalize_cache_report(cache_report),
            "pages": page_debug,
        },
    )


def _convert_pdf_with_pypdf(path: Path, options: ConversionOptions) -> ConversionResult:
    try:
        from pypdf import PdfReader
    except ImportError as exc:
        raise ConversionError("PDF conversion requires the 'pypdf' package.") from exc

    reader = PdfReader(str(path))
    pages = []
    markdown_sections = []
    page_debug: list[dict[str, Any]] = []
    selected_pages = options.page_numbers or set(range(1, len(reader.pages) + 1))
    total_selected = len(selected_pages)
    processed = 0
    for index, page in enumerate(reader.pages, start=1):
        if index not in selected_pages:
            continue
        plain_text = (page.extract_text() or "").strip()
        layout_text = (page.extract_text(extraction_mode="layout") or plain_text).strip()
        pages.append((index, plain_text, layout_text))
        markdown_sections.append(_render_pdf_page_markdown(index, layout_text or plain_text))
        debug = {
            "page": index,
            "layout": "text",
            "layout_confidence": 0.55,
            "text_blocks": 0,
            "image_blocks": 0,
            "drawing_blocks": 0,
            "table_count": 0,
            "text_line_count": 0,
            "text_char_count": len(plain_text),
            "dominant_signal": "text",
            "signal_scores": {"text": 0.55, "table": 0.0, "diagram": 0.0},
            "diagram": {"boxes": [], "connectors": 0, "connector_segments": [], "edges": [], "unlabeled_boxes": 0},
            "tables": [],
            "ocr_status": "processed",
            "ocr_confidence": {},
            "text_length": len(plain_text),
            "markdown_length": len(markdown_sections[-1]),
        }
        page_debug.append(debug)
        processed += 1
        if options.progress_callback:
            options.progress_callback(index, len(reader.pages), processed, total_selected, "processed")
        if options.analysis_callback:
            options.analysis_callback(debug)

    text = "\n\n".join(
        f"[Page {index}]\n{content}" if content else f"[Page {index}]"
        for index, content, _layout in pages
    ).strip()
    html_sections = [_render_pdf_page_html(index, layout_text or plain_text) for index, plain_text, layout_text in pages]
    html = _build_html_document(path.name, "\n".join(html_sections) or "<p></p>")
    markdown = "\n\n".join(section for section in markdown_sections if section).strip()
    return ConversionResult(
        source=path,
        text=text,
        html=html,
        markdown=markdown,
        debug_data={
            "kind": "pdf",
            "engine": "pypdf",
            "page_count": len(reader.pages),
            "selected_pages": sorted(selected_pages),
            "pages": page_debug,
        },
    )


def _convert_xlsx(path: Path) -> ConversionResult:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ConversionError("XLSX conversion requires the 'openpyxl' package.") from exc

    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        text_parts = []
        html_parts = []
        markdown_parts = []
        for sheet in workbook.worksheets:
            rows = [[_format_cell(value) for value in row] for row in sheet.iter_rows(values_only=True)]
            text_parts.append(_render_sheet_text(sheet.title, rows))
            html_parts.append(_render_sheet_html(sheet.title, rows))
            markdown_parts.append(_render_sheet_markdown(sheet.title, rows))
    finally:
        workbook.close()

    html = _build_html_document(path.name, "\n".join(html_parts))
    return ConversionResult(
        source=path,
        text="\n\n".join(text_parts).strip(),
        html=html,
        markdown="\n\n".join(part for part in markdown_parts if part).strip(),
        debug_data={"kind": "xlsx", "sheet_count": len(workbook.worksheets)},
    )


def _convert_xls(path: Path) -> ConversionResult:
    try:
        import xlrd
    except ImportError as exc:
        raise ConversionError("XLS conversion requires the 'xlrd' package.") from exc

    book = xlrd.open_workbook(path)
    text_parts = []
    html_parts = []
    markdown_parts = []
    for sheet in book.sheets():
        rows = []
        for row_index in range(sheet.nrows):
            rows.append([_format_cell(sheet.cell_value(row_index, col)) for col in range(sheet.ncols)])
        text_parts.append(_render_sheet_text(sheet.name, rows))
        html_parts.append(_render_sheet_html(sheet.name, rows))
        markdown_parts.append(_render_sheet_markdown(sheet.name, rows))

    html = _build_html_document(path.name, "\n".join(html_parts))
    return ConversionResult(
        source=path,
        text="\n\n".join(text_parts).strip(),
        html=html,
        markdown="\n\n".join(part for part in markdown_parts if part).strip(),
        debug_data={"kind": "xls", "sheet_count": len(book.sheets())},
    )


def _convert_plain_text(path: Path) -> ConversionResult:
    text = path.read_text(encoding="utf-8", errors="replace").strip()
    html = _build_html_document(path.name, f"<pre>{escape(text)}</pre>")
    return ConversionResult(source=path, text=text, html=html, debug_data={"kind": "text", "text_length": len(text)})


def _convert_html(path: Path) -> ConversionResult:
    raw_html = path.read_text(encoding="utf-8", errors="replace")
    text = _strip_html_tags(raw_html)
    return ConversionResult(source=path, text=text, html=raw_html, debug_data={"kind": "html", "text_length": len(text)})


def _render_sheet_text(title: str, rows: list[list[str]]) -> str:
    lines = [f"[Sheet] {title}"]
    for row in rows:
        if any(cell for cell in row):
            lines.append("\t".join(row))
    return "\n".join(lines).strip()


def _render_sheet_html(title: str, rows: list[list[str]]) -> str:
    rendered_rows = []
    for row in rows:
        cells = "".join(f"<td>{escape(cell)}</td>" for cell in row)
        rendered_rows.append(f"<tr>{cells}</tr>")
    table = "<table>\n" + "\n".join(rendered_rows) + "\n</table>"
    return f"<section><h2>{escape(title)}</h2>{table}</section>"


def _render_sheet_markdown(title: str, rows: list[list[str]]) -> str:
    cleaned_rows = [row for row in rows if any(cell.strip() for cell in row)]
    parts = [f"## {title}"]
    if cleaned_rows:
        parts.append(_render_markdown_table(cleaned_rows))
    return "\n\n".join(part for part in parts if part).strip()


def _format_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _build_html_document(title: str, body: str, warnings: list[str] | None = None) -> str:
    warning_html = ""
    if warnings:
        items = "".join(f"<li>{escape(item)}</li>" for item in warnings)
        warning_html = f"<aside><h2>Warnings</h2><ul>{items}</ul></aside>"

    return (
        "<!doctype html>\n"
        "<html lang=\"en\">\n"
        "<head>\n"
        "  <meta charset=\"utf-8\">\n"
        f"  <title>{escape(title)}</title>\n"
        "  <style>\n"
        "    body { font-family: sans-serif; line-height: 1.6; margin: 2rem; color: #1f2937; }\n"
        "    h1, h2, h3, h4 { line-height: 1.3; }\n"
        "    .pdf-page { border-top: 1px solid #d1d5db; margin-top: 2rem; padding-top: 1.5rem; }\n"
        "    .pdf-page:first-of-type { border-top: 0; margin-top: 0; padding-top: 0; }\n"
        "    .pdf-page-frame { position: relative; margin: 1rem 0 1.5rem; border: 1px solid #cbd5e1; background: #fff; overflow: hidden; }\n"
        "    .pdf-page-bg { display: block; width: 100%; height: auto; }\n"
        "    .pdf-text-layer { position: absolute; inset: 0; color: transparent; user-select: text; }\n"
        "    .pdf-text-span { position: absolute; white-space: pre; transform-origin: top left; line-height: 1; color: transparent; }\n"
        "    .pdf-debug-layer { position: absolute; inset: 0; pointer-events: none; }\n"
        "    .pdf-debug-box { position: absolute; border: 2px solid rgba(220, 38, 38, 0.9); background: rgba(220, 38, 38, 0.08); box-sizing: border-box; }\n"
        "    .pdf-debug-box.inferred { border-color: rgba(37, 99, 235, 0.9); background: rgba(37, 99, 235, 0.08); }\n"
        "    .pdf-debug-label { position: absolute; top: -1.4rem; left: 0; padding: 0.1rem 0.35rem; font: 600 12px/1.2 sans-serif; color: #111827; background: rgba(255, 255, 255, 0.92); border: 1px solid rgba(15, 23, 42, 0.2); white-space: nowrap; }\n"
        "    .pdf-debug-badge { position: absolute; top: 0.5rem; right: 0.5rem; padding: 0.2rem 0.45rem; font: 600 12px/1.2 sans-serif; color: #fff; background: rgba(217, 119, 6, 0.9); border-radius: 999px; }\n"
        "    .pdf-preview { margin: 1rem 0 1.5rem; border: 1px solid #d1d5db; background: #f8fafc; padding: 0.75rem; }\n"
        "    .pdf-preview img { display: block; width: 100%; height: auto; }\n"
        "    .pdf-reconstructed { margin: 1rem 0 1.5rem; border: 1px solid #d1d5db; background: #fff; padding: 1rem; }\n"
        "    .pdf-reconstructed-page { display: flex; flex-direction: column; gap: 0.75rem; margin: 0; }\n"
        "    .pdf-reconstructed-page.pdf-reconstructed-form { gap: 0.9rem; }\n"
        "    .pdf-reconstructed-row { display: grid; grid-template-columns: repeat(12, minmax(0, 1fr)); column-gap: 1rem; align-items: start; margin: 0; padding: 0; border: 0; min-inline-size: 0; }\n"
        "    .pdf-reconstructed-row.columns-2 { column-gap: 1.5rem; }\n"
        "    .pdf-reconstructed-row.columns-3 { column-gap: 1.25rem; }\n"
        "    .pdf-form-title { margin: 0 0 0.4rem; text-align: center; }\n"
        "    .pdf-form-title h1 { margin: 0; font-size: 2rem; letter-spacing: 0.08em; }\n"
        "    .pdf-form-overview, .pdf-form-section { margin: 0; }\n"
        "    .pdf-form-section > header { display: flex; align-items: flex-end; justify-content: flex-start; margin: 0 0 0.35rem; }\n"
        "    .pdf-form-section > header h2 { margin: 0; font-size: 1.1rem; letter-spacing: 0.04em; }\n"
        "    .pdf-form-section table { margin-top: 0; }\n"
        "    .pdf-reconstructed-summary { margin-bottom: 0.85rem; }\n"
        "    .pdf-reconstructed-summary p { margin: 0; color: #64748b; font-size: 0.92rem; }\n"
        "    .pdf-reconstructed-item { min-width: 0; }\n"
        "    .pdf-reconstructed-item.centered { text-align: center; }\n"
        "    .pdf-reconstructed-item.full-width > *:first-child { margin-top: 0; }\n"
        "    .pdf-reconstructed-item.callout { padding: 0.75rem 0.9rem; border: 1px solid #cbd5e1; border-radius: 0.5rem; background: #f8fafc; }\n"
        "    .pdf-reconstructed-item.narrow { font-size: 0.96rem; }\n"
        "    .pdf-reconstructed-item.table { overflow-x: auto; }\n"
        "    .pdf-reconstructed-item.form-field { padding: 0.15rem 0; }\n"
        "    .pdf-reconstructed-item.form-field p { margin-bottom: 0.25rem; }\n"
        "    .pdf-semantic { margin-top: 1rem; }\n"
        "    .pdf-semantic details { margin-top: 1rem; }\n"
        "    .pdf-semantic summary { cursor: pointer; color: #334155; font-weight: 600; }\n"
        "    .pdf-table-wrap { overflow-x: auto; }\n"
        "    table { border-collapse: collapse; margin: 1rem 0; width: 100%; }\n"
        "    td, th { border: 1px solid #ccc; padding: 0.35rem 0.5rem; vertical-align: top; }\n"
        "    th { background: #f3f4f6; }\n"
        "    pre { white-space: pre-wrap; }\n"
        "    ul, ol { padding-left: 1.4rem; }\n"
        "    p { margin: 0.4rem 0 0.8rem; }\n"
        "    aside { background: #fff8dc; border: 1px solid #e0c97f; padding: 1rem; }\n"
        "  </style>\n"
        "</head>\n"
        "<body>\n"
        f"<h1>{escape(title)}</h1>\n"
        f"{warning_html}\n"
        f"{body}\n"
        "</body>\n"
        "</html>\n"
    )


def _strip_html_tags(raw_html: str) -> str:
    try:
        from html.parser import HTMLParser
    except ImportError:
        return raw_html

    class Stripper(HTMLParser):
        def __init__(self) -> None:
            super().__init__()
            self.parts: list[str] = []

        def handle_data(self, data: str) -> None:
            if data.strip():
                self.parts.append(data)

    parser = Stripper()
    parser.feed(raw_html)
    return "\n".join(parser.parts).strip()


def _markdown_from_html(raw_html: str) -> str:
    try:
        from html.parser import HTMLParser
    except ImportError:
        return ""

    class MarkdownParser(HTMLParser):
        def __init__(self) -> None:
            super().__init__()
            self.parts: list[str] = []
            self.block_stack: list[dict[str, Any]] = []
            self.skip_depth = 0
            self.list_depth = 0
            self.current_table: list[list[str]] | None = None
            self.current_row: list[str] | None = None
            self.current_cell: list[str] | None = None

        def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
            attrs_map = dict(attrs)
            classes = set((attrs_map.get("class") or "").split())
            should_skip = tag in {"head", "style", "script", "img"} or bool(
                {"pdf-page-frame", "pdf-preview", "pdf-text-layer"} & classes
            )
            if self.skip_depth or should_skip:
                self.skip_depth += 1
                return

            if tag in {"ul", "ol"}:
                self.list_depth += 1
                return
            if tag == "br":
                self._append_text("\n")
                return
            if tag == "table":
                self.current_table = []
                return
            if tag == "tr":
                self.current_row = []
                return
            if tag in {"td", "th"}:
                self.current_cell = []
                return
            if tag in {"p", "pre", "li", "summary", "h1", "h2", "h3", "h4", "h5", "h6"}:
                self.block_stack.append({"tag": tag, "chunks": []})

        def handle_endtag(self, tag: str) -> None:
            if self.skip_depth:
                self.skip_depth -= 1
                return

            if tag in {"ul", "ol"}:
                self.list_depth = max(self.list_depth - 1, 0)
                self._append_part("\n")
                return
            if tag in {"td", "th"}:
                if self.current_row is not None:
                    self.current_row.append(_normalize_markdown_text("".join(self.current_cell or [])))
                self.current_cell = None
                return
            if tag == "tr":
                if self.current_table is not None and self.current_row:
                    self.current_table.append(self.current_row)
                self.current_row = None
                return
            if tag == "table":
                if self.current_table:
                    self._append_part(_render_markdown_table(self.current_table) + "\n\n")
                self.current_table = None
                return
            if tag in {"p", "pre", "li", "summary", "h1", "h2", "h3", "h4", "h5", "h6"}:
                if not self.block_stack:
                    return
                block = self.block_stack.pop()
                text = "".join(block["chunks"])
                self._emit_block(block["tag"], text)

        def handle_data(self, data: str) -> None:
            if self.skip_depth or not data:
                return
            self._append_text(data)

        def _append_text(self, text: str) -> None:
            if self.current_cell is not None:
                self.current_cell.append(text)
                return
            if self.block_stack:
                self.block_stack[-1]["chunks"].append(text)

        def _append_part(self, text: str) -> None:
            if text:
                self.parts.append(text)

        def _emit_block(self, tag: str, text: str) -> None:
            preserve_newlines = tag == "pre"
            cleaned = _normalize_markdown_text(text, preserve_newlines=preserve_newlines)
            if not cleaned:
                return
            if tag.startswith("h") and len(tag) == 2 and tag[1].isdigit():
                level = min(max(int(tag[1]), 1), 6)
                self._append_part(f"{'#' * level} {cleaned}\n\n")
                return
            if tag == "li":
                self._append_part(f"- {cleaned}\n")
                return
            if tag == "pre":
                self._append_part(f"```\n{cleaned}\n```\n\n")
                return
            self._append_part(f"{cleaned}\n\n")

    parser = MarkdownParser()
    parser.feed(raw_html)
    parser.close()
    return _tidy_markdown("".join(parser.parts))


def _normalize_markdown_text(text: str, preserve_newlines: bool = False) -> str:
    text = text.replace("\xa0", " ")
    if preserve_newlines:
        lines = [re.sub(r"[ \t]+", " ", line).rstrip() for line in text.splitlines()]
        return "\n".join(line for line in lines if line.strip()).strip()
    text = re.sub(r"\s*\n\s*", "\n", text)
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{2,}", "\n", text)
    return text.strip()


def _render_markdown_table(rows: list[list[str]]) -> str:
    normalized = _normalize_table_rows(rows)
    escaped_rows = [[cell.replace("|", "\\|").replace("\n", "<br />").strip() for cell in row] for row in normalized]
    if not escaped_rows:
        return ""

    header_rows = _detect_table_header_rows(normalized)
    if header_rows:
        header = _merge_table_header_band(escaped_rows[:header_rows])
        body = escaped_rows[header_rows:]
    else:
        header = [f"Column {index + 1}" for index in range(len(escaped_rows[0]))]
        body = escaped_rows
    if not body:
        body = [[""] * len(header)]
    separator = ["---"] * len(header)
    lines = [
        "| " + " | ".join(header) + " |",
        "| " + " | ".join(separator) + " |",
    ]
    lines.extend("| " + " | ".join(row) + " |" for row in body)
    return "\n".join(lines)


def _tidy_markdown(markdown: str) -> str:
    markdown = re.sub(r"\n{3,}", "\n\n", markdown)
    return markdown.strip()


def _render_ocr_markdown(text: str) -> str:
    lines = _normalize_pdf_lines(text)
    if not lines:
        return _normalize_markdown_text(text, preserve_newlines=True)
    blocks = _group_pdf_blocks(lines)
    rendered = [_render_pdf_block_markdown(block_type, block_lines) for block_type, block_lines in blocks]
    body = "\n\n".join(part for part in rendered if part).strip()
    if body:
        return body
    return _normalize_markdown_text(text, preserve_newlines=True)


def _extract_pymupdf_page(
    page: Any,
    index: int,
    options: ConversionOptions,
    source_path: Path,
    document: Any,
    cache_report: dict[str, dict[str, Any]],
    cached_ocr_payload: dict[str, str] | None = None,
) -> tuple[str, str, str, str, dict[str, Any]]:
    page_dict = page.get_text("dict", sort=True)
    diagram = _extract_pymupdf_diagram_primitives(page, page_dict)
    tables, table_cache_status = _extract_pymupdf_tables(
        page,
        page_dict,
        source_path,
        document,
        index,
        options.cache_dir,
        cache_report,
    )
    blocks = page_dict.get("blocks", [])
    analysis = _analyze_pdf_page(blocks, len(tables))
    table_bboxes = [table["bbox"] for table in tables]

    text_parts: list[str] = []
    html_parts: list[str] = []
    markdown_parts: list[str] = []
    pending_table_index = 0

    for block in blocks:
        bbox = tuple(block.get("bbox", (0, 0, 0, 0)))

        while pending_table_index < len(tables) and tables[pending_table_index]["bbox"][1] < bbox[1]:
            table = tables[pending_table_index]
            html_parts.append(table["html"])
            text_parts.append(table["text"])
            markdown_parts.append(table["markdown"])
            pending_table_index += 1

        if block.get("type") != 0:
            continue
        if any(_bbox_overlaps(bbox, table_bbox) for table_bbox in table_bboxes):
            continue

        block_text, block_html, block_markdown = _render_pymupdf_text_block(block)
        if block_text:
            text_parts.append(block_text)
        if block_html:
            html_parts.append(block_html)
        if block_markdown:
            markdown_parts.append(block_markdown)

    while pending_table_index < len(tables):
        table = tables[pending_table_index]
        html_parts.append(table["html"])
        text_parts.append(table["text"])
        markdown_parts.append(table["markdown"])
        pending_table_index += 1

    visible_text = "\n".join(part for part in text_parts if part).strip()
    ocr_payload = {
        "text": "",
        "overlay": "",
        "profile": "",
        "strategy": {},
        "words": [],
        "image_width": 0,
        "image_height": 0,
    }
    status = "native-text"
    if options.ocr_mode != "off" and (options.ocr_mode == "force" or not visible_text):
        if cached_ocr_payload:
            ocr_payload = cached_ocr_payload
        else:
            ocr_payload = _extract_pdf_page_ocr_payload(
                page,
                source_path,
                document,
                index,
                cache_dir=options.cache_dir,
                cache_report=cache_report,
                force=options.ocr_mode == "force",
                dpi=options.ocr_dpi,
                layout_hint=analysis.layout,
            )
        status = ocr_payload.get("status", "ocr")
        if ocr_payload["text"]:
            visible_text = ocr_payload["text"] if not visible_text else f"{visible_text}\n{ocr_payload['text']}".strip()
            html_parts.append(f"<pre>{escape(ocr_payload['text'])}</pre>")
            markdown_parts.append(_render_ocr_markdown(ocr_payload["text"]))
            diagram = _merge_ocr_labels_into_diagram(diagram, ocr_payload, page)

    body = "\n".join(part for part in html_parts if part) or "<p></p>"
    if diagram["boxes"] or diagram.get("edges"):
        body += _render_diagram_summary_html(diagram)
    markdown_body = "\n\n".join(part for part in markdown_parts if part).strip()
    if diagram["boxes"] or diagram.get("edges"):
        markdown_body = "\n\n".join(part for part in [markdown_body, _render_diagram_summary_markdown(diagram)] if part).strip()
    reconstructed_html = _render_pdf_page_reconstructed(
        page,
        blocks,
        tables,
        table_bboxes,
        body if body != "<p></p>" else "",
        ocr_payload,
        diagram,
    )
    preview_html = _render_pdf_page_preview(
        page,
        source_path,
        document,
        index,
        options.cache_dir,
        cache_report,
    )
    faithful_html = _render_pdf_page_faithful(
        page,
        page_dict,
        ocr_payload["overlay"],
        source_path=source_path,
        document=document,
        page_index=index,
        cache_dir=options.cache_dir,
        cache_report=cache_report,
        tables=tables,
        ocr_status=status,
        debug_overlays=options.debug_overlays,
    )
    semantic_html = f"<div class=\"pdf-semantic\">{body}</div>"
    if body and body != "<p></p>":
        semantic_html = (
            "<div class=\"pdf-semantic\"><details>"
            "<summary>Structured Content</summary>"
            f"{body}</details></div>"
        )
    page_html = faithful_html or preview_html
    if options.pdf_mode == "semantic":
        page_html = ""
    elif options.pdf_mode == "reconstructed":
        page_html = reconstructed_html
        semantic_html = ""
    elif options.pdf_mode == "faithful":
        semantic_html = ""
    page_meta = (
        f"<p><small>Layout: {analysis.layout} | text blocks: {analysis.text_blocks} | "
        f"tables: {analysis.table_count} | images: {analysis.image_blocks} | drawings: {analysis.drawing_blocks}</small></p>"
    )
    if options.pdf_mode == "faithful":
        page_meta = ""
    page_markdown = f"## Page {index}"
    if markdown_body:
        page_markdown = f"{page_markdown}\n\n{markdown_body}"
    debug = {
        "page": index,
        "layout": analysis.layout,
        "layout_confidence": analysis.layout_confidence,
        "text_blocks": analysis.text_blocks,
        "image_blocks": analysis.image_blocks,
        "drawing_blocks": analysis.drawing_blocks,
        "table_count": analysis.table_count,
        "text_line_count": analysis.text_line_count,
        "text_char_count": analysis.text_char_count,
        "dominant_signal": analysis.dominant_signal,
        "signal_scores": analysis.signal_scores,
        "diagram": diagram,
        "tables": [
            {
                "bbox": [round(value, 2) for value in table["bbox"]],
                "confidence": table.get("confidence"),
                "source": table.get("source", "unknown"),
            }
            for table in tables
        ],
        "ocr_status": status,
        "ocr_used": status != "native-text",
        "ocr_profile": ocr_payload.get("profile", ""),
        "ocr_strategy": ocr_payload.get("strategy", {}),
        "ocr_confidence": ocr_payload.get("confidence_summary", {}),
        "cache": {
            "tables": table_cache_status,
            "ocr": "hit" if status == "cache-hit" else ("miss" if status.startswith("ocr-") else "unused"),
        },
        "text_length": len(visible_text),
        "markdown_length": len(page_markdown),
    }
    return visible_text, (
        f"<section class=\"pdf-page pdf-layout-{analysis.layout}\"><h2>Page {index}</h2>"
        f"{page_html}{page_meta}{semantic_html}</section>"
    ), page_markdown, status, debug


def _precompute_document_ocr(
    document: Any,
    source_path: Path,
    selected_pages: set[int],
    options: ConversionOptions,
    cache_report: dict[str, dict[str, Any]],
) -> dict[int, dict[str, str]]:
    if options.ocr_mode != "force":
        return {}

    tasks: list[tuple[int, bytes, bool, int, str, dict[str, Any], tuple[float, float]]] = []
    payloads: dict[int, dict[str, str]] = {}
    for index, page in enumerate(document, start=1):
        if index not in selected_pages:
            continue
        page_dict = page.get_text("dict", sort=True)
        layout_hint = _analyze_pdf_page(page_dict.get("blocks", []), 0).layout
        profile = _build_ocr_profile(layout_hint, force=True)
        diagram = _extract_pymupdf_diagram_primitives(page, page_dict)
        profile_token = _ocr_profile_cache_token(profile)
        cache_key = _build_page_cache_key("ocr", source_path, index, dpi=options.ocr_dpi, variant="force", profile=profile_token)
        cache_meta = _build_page_cache_metadata(source_path, document, index, dpi=options.ocr_dpi, variant="force", profile=profile_token)
        cached = _load_ocr_cache(options.cache_dir, cache_key, cache_meta, cache_report)
        if cached:
            cached["status"] = "cache-hit"
            payloads[index] = cached
            continue
        raster = _get_cached_page_raster(
            page,
            source_path,
            document,
            index,
            options.ocr_dpi,
            options.cache_dir,
            cache_report,
        )
        if not raster:
            continue
        tasks.append(
            (
                index,
                base64.b64decode(raster["data"]),
                True,
                options.ocr_dpi,
                profile["key"],
                diagram,
                (float(page.rect.width or 0.0), float(page.rect.height or 0.0)),
            )
        )

    if not tasks:
        return payloads

    workers = max(1, min(options.ocr_workers, os.cpu_count() or 1))
    if workers <= 1:
        results = map(_extract_ocr_from_raster, tasks)
    else:
        executor = ThreadPoolExecutor(max_workers=workers)
        try:
            results = executor.map(_extract_ocr_from_raster, tasks)
            for index, payload in results:
                payload["status"] = "ocr-force"
                payloads[index] = payload
                profile_key = str(payload.get("profile", "force-text"))
                profile_token = _ocr_profile_cache_token(_build_ocr_profile(profile_key.split("-", 1)[1], force=True))
                cache_key = _build_page_cache_key("ocr", source_path, index, dpi=options.ocr_dpi, variant="force", profile=profile_token)
                cache_meta = _build_page_cache_metadata(source_path, document, index, dpi=options.ocr_dpi, variant="force", profile=profile_token)
                _save_ocr_cache(options.cache_dir, cache_key, cache_meta, payload, cache_report)
        finally:
            executor.shutdown(wait=True)
        return payloads

    for index, payload in results:
        payload["status"] = "ocr-force"
        payloads[index] = payload
        profile_key = str(payload.get("profile", "force-text"))
        profile_token = _ocr_profile_cache_token(_build_ocr_profile(profile_key.split("-", 1)[1], force=True))
        cache_key = _build_page_cache_key("ocr", source_path, index, dpi=options.ocr_dpi, variant="force", profile=profile_token)
        cache_meta = _build_page_cache_metadata(source_path, document, index, dpi=options.ocr_dpi, variant="force", profile=profile_token)
        _save_ocr_cache(options.cache_dir, cache_key, cache_meta, payload, cache_report)
    return payloads


def _extract_pymupdf_tables(
    page: Any,
    page_dict: dict[str, Any] | None = None,
    source_path: Path | None = None,
    document: Any | None = None,
    page_index: int | None = None,
    cache_dir: Path | None = None,
    cache_report: dict[str, dict[str, Any]] | None = None,
) -> tuple[list[dict[str, Any]], str]:
    if source_path and document is not None and page_index is not None:
        cache_key = _build_page_cache_key("tables", source_path, page_index)
        cache_meta = _build_page_cache_metadata(source_path, document, page_index)
        cached = _load_json_cache(cache_dir, "tables", cache_key, cache_meta, cache_report)
        if isinstance(cached, list):
            return [_normalize_cached_table_entry(item) for item in cached], "hit"
        cache_path = _cache_file_path(cache_dir, "tables", cache_key)
        with _cache_lock(cache_path):
            if cache_path and cache_path.exists():
                cached = _load_json_cache_from_path(cache_path, "tables", cache_meta, cache_report)
                if isinstance(cached, list):
                    return [_normalize_cached_table_entry(item) for item in cached], "hit"

            try:
                with redirect_stdout(io.StringIO()), redirect_stderr(io.StringIO()):
                    finder = page.find_tables()
            except Exception:
                finder = None

            tables = []
            native_bboxes: list[tuple[float, float, float, float]] = []
            for table in finder.tables if finder else []:
                rows = table.extract() or []
                cleaned_rows = [
                    [_clean_pdf_text(cell or "") for cell in row]
                    for row in rows
                    if any((cell or "").strip() for cell in row)
                ]
                if not cleaned_rows:
                    continue

                bbox = tuple(table.bbox)
                native_bboxes.append(bbox)
                html, text, markdown = _render_table_rows(cleaned_rows, 0.98, inferred=False)
                tables.append(
                    {
                        "bbox": bbox,
                        "html": html,
                        "text": text,
                        "markdown": markdown,
                        "confidence": 0.98,
                        "source": "native",
                        "column_signatures": [],
                    }
                )

            if page_dict:
                tables.extend(_infer_pymupdf_tables(page_dict.get("blocks", []), native_bboxes))

            tables.sort(key=lambda item: item["bbox"][1])
            _save_json_cache(cache_dir, "tables", cache_key, cache_meta, tables, cache_report)
            return tables, "miss"

    try:
        with redirect_stdout(io.StringIO()), redirect_stderr(io.StringIO()):
            finder = page.find_tables()
    except Exception:
        finder = None

    tables = []
    native_bboxes: list[tuple[float, float, float, float]] = []
    for table in finder.tables if finder else []:
        rows = table.extract() or []
        cleaned_rows = [
            [_clean_pdf_text(cell or "") for cell in row]
            for row in rows
            if any((cell or "").strip() for cell in row)
        ]
        if not cleaned_rows:
            continue

        bbox = tuple(table.bbox)
        native_bboxes.append(bbox)
        html, text, markdown = _render_table_rows(cleaned_rows, 0.98, inferred=False)
        tables.append(
            {
                "bbox": bbox,
                "html": html,
                "text": text,
                "markdown": markdown,
                "confidence": 0.98,
                "source": "native",
                "column_signatures": [],
            }
        )

    if page_dict:
        tables.extend(_infer_pymupdf_tables(page_dict.get("blocks", []), native_bboxes))

    tables.sort(key=lambda item: item["bbox"][1])
    if source_path and document is not None and page_index is not None:
        cache_key = _build_page_cache_key("tables", source_path, page_index)
        cache_meta = _build_page_cache_metadata(source_path, document, page_index)
        _save_json_cache(cache_dir, "tables", cache_key, cache_meta, tables, cache_report)
    return tables, "miss"


def _infer_pymupdf_tables(
    blocks: list[dict[str, Any]],
    occupied_bboxes: list[tuple[float, float, float, float]],
) -> list[dict[str, Any]]:
    line_rows = _collect_inferred_table_rows(blocks, occupied_bboxes)
    if not line_rows:
        return []

    tables: list[dict[str, Any]] = []
    current: dict[str, Any] | None = None

    for row in line_rows:
        if current:
            belongs, merged_into_previous = _row_belongs_to_inferred_table(current, row)
            if belongs:
                if not merged_into_previous:
                    current["rows"].append(row)
                current["match_scores"].append(row["match_score"])
                current["bbox"] = _merge_bboxes(current["bbox"], row["bbox"])
                continue
        finalized = _finalize_inferred_table(current)
        if finalized:
            tables.append(finalized)
        current = {
            "rows": [row],
            "column_signatures": list(row["column_signatures"]),
            "match_scores": [row["match_score"]],
            "bbox": row["bbox"],
        }

    finalized = _finalize_inferred_table(current)
    if finalized:
        tables.append(finalized)

    return _merge_adjacent_inferred_tables(tables)


def _collect_inferred_table_rows(
    blocks: list[dict[str, Any]],
    occupied_bboxes: list[tuple[float, float, float, float]],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for block in blocks:
        if block.get("type") != 0:
            continue
        for line in block.get("lines", []):
            spans = [span for span in line.get("spans", []) if (span.get("text") or "").strip()]
            groups = _group_pymupdf_line_spans(spans)
            if len(groups) < 2 or len(groups) > 6:
                continue
            bbox = _line_bbox(spans)
            if any(_bbox_overlaps(bbox, occupied) for occupied in occupied_bboxes):
                continue
            row = _build_inferred_table_row(groups, bbox)
            if row:
                rows.append(row)
    rows.sort(key=lambda item: (item["bbox"][1], item["bbox"][0]))
    return rows


def _build_inferred_table_row(
    groups: list[tuple[float, float, list[str]]],
    bbox: tuple[float, float, float, float],
) -> dict[str, Any] | None:
    row = [" ".join(texts).strip() for _x0, _x1, texts in groups]
    compact = [cell for cell in row if cell]
    if len(compact) < 2:
        return None
    if all(_looks_like_pdf_heading(cell) for cell in compact):
        return None
    signatures = [(x0 + x1) / 2 for x0, x1, _texts in groups]
    return {
        "cells": row,
        "groups": groups,
        "bbox": bbox,
        "column_signatures": signatures,
        "match_score": 0.7,
    }


def _row_belongs_to_inferred_table(table: dict[str, Any], row: dict[str, Any]) -> tuple[bool, bool]:
    last_row = table["rows"][-1]
    gap = row["bbox"][1] - last_row["bbox"][3]
    last_height = max(last_row["bbox"][3] - last_row["bbox"][1], 1.0)
    row_height = max(row["bbox"][3] - row["bbox"][1], 1.0)
    if gap > max(36.0, max(last_height, row_height) * 1.7):
        return False, False

    aligned, match_ratio = _align_groups_to_columns(
        row["groups"],
        table["column_signatures"],
        allow_expand=len(table["column_signatures"]) < 6,
    )
    if not aligned or match_ratio < 0.55:
        return False, False

    row["cells"] = aligned
    row["match_score"] = match_ratio
    if _should_merge_inferred_row(last_row, row):
        last_row["cells"] = _merge_row_cells(last_row["cells"], row["cells"])
        last_row["bbox"] = _merge_bboxes(last_row["bbox"], row["bbox"])
        return True, True
    return True, False


def _finalize_inferred_table(table: dict[str, Any] | None) -> dict[str, Any] | None:
    if not table:
        return None

    rows = [row["cells"] for row in table["rows"] if any(cell.strip() for cell in row["cells"])]
    if len(rows) < 2:
        return None
    widths = {len(row) for row in rows}
    if len(widths) != 1 or next(iter(widths)) < 2:
        return None

    confidence = _score_inferred_table(rows, table["match_scores"])
    if confidence < 0.58:
        return None
    html, text, markdown = _render_table_rows(rows, confidence, inferred=True)
    return {
        "bbox": table["bbox"],
        "html": html,
        "text": text,
        "markdown": markdown,
        "confidence": confidence,
        "source": "inferred",
        "column_signatures": list(table["column_signatures"]),
    }


def _merge_adjacent_inferred_tables(tables: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if not tables:
        return []

    merged: list[dict[str, Any]] = []
    current = tables[0]
    for table in tables[1:]:
        if (
            current.get("source") == "inferred"
            and table.get("source") == "inferred"
            and _table_columns_are_compatible(current.get("column_signatures", []), table.get("column_signatures", []))
            and table["bbox"][1] - current["bbox"][3] <= 28.0
        ):
            current_rows = [line.split("\t") for line in current["text"].splitlines() if line.strip()]
            next_rows = [line.split("\t") for line in table["text"].splitlines() if line.strip()]
            combined_rows = current_rows + next_rows
            confidence = min(max(current.get("confidence", 0.0), table.get("confidence", 0.0)) + 0.04, 0.97)
            html, text, markdown = _render_table_rows(combined_rows, confidence, inferred=True)
            current = {
                "bbox": _merge_bboxes(current["bbox"], table["bbox"]),
                "html": html,
                "text": text,
                "markdown": markdown,
                "confidence": confidence,
                "source": "inferred",
                "column_signatures": current.get("column_signatures", table.get("column_signatures", [])),
            }
            continue
        merged.append(current)
        current = table

    merged.append(current)
    return merged


def _score_inferred_table(rows: list[list[str]], match_scores: list[float]) -> float:
    row_density = sum(sum(1 for cell in row if cell.strip()) / max(len(row), 1) for row in rows) / max(len(rows), 1)
    avg_match = sum(match_scores) / max(len(match_scores), 1)
    row_bonus = min(len(rows) / 4, 1.0) * 0.12
    confidence = 0.38 + (avg_match * 0.34) + (row_density * 0.16) + row_bonus
    return round(min(confidence, 0.97), 2)


def _render_table_rows(rows: list[list[str]], confidence: float | None, inferred: bool) -> tuple[str, str, str]:
    cleaned_rows = [[_clean_table_cell(cell) for cell in row] for row in rows]
    normalized_rows = _normalize_table_rows(cleaned_rows)
    header_rows = _detect_table_header_rows(normalized_rows)
    header = _merge_table_header_band(normalized_rows[:header_rows]) if header_rows else None
    body_rows = normalized_rows[header_rows:] if header_rows else normalized_rows
    text = "\n".join("\t".join(cell.replace("\n", " / ").strip() for cell in row) for row in normalized_rows)

    table_parts: list[str] = []
    if header:
        header_html = "".join(f"<th>{_render_table_cell_html(cell)}</th>" for cell in header)
        table_parts.append(f"<thead><tr>{header_html}</tr></thead>")
    body_html = []
    for row in body_rows:
        body_html.append("<tr>" + "".join(f"<td>{_render_table_cell_html(cell)}</td>" for cell in row) + "</tr>")
    if body_html:
        table_parts.append("<tbody>\n" + "\n".join(body_html) + "\n</tbody>")

    meta = ""
    markdown_meta = ""
    if confidence is not None:
        label = "Inferred table" if inferred else "Detected table"
        meta = f"<p><small>{label} | confidence: {confidence:.2f}</small></p>"
        markdown_meta = f"> {label} | confidence: {confidence:.2f}\n\n"
    html = meta + "<div class=\"pdf-table-wrap\"><table>\n" + "\n".join(table_parts) + "\n</table></div>"
    markdown = markdown_meta + _render_markdown_table(normalized_rows)
    return html, text, markdown


def _normalize_table_rows(rows: list[list[str]]) -> list[list[str]]:
    width = max((len(row) for row in rows), default=0)
    return [row + [""] * max(0, width - len(row)) for row in rows]


def _render_table_cell_html(cell: str) -> str:
    return "<br />".join(escape(part) for part in cell.splitlines()) if cell else ""


def _clean_table_cell(cell: str) -> str:
    cleaned = _clean_pdf_text(cell)
    if not cleaned:
        return ""

    lines = [line.strip() for line in cleaned.splitlines() if line.strip()]
    if len(lines) >= 2 and _looks_like_trailing_artifact(lines[-1], lines[-2]):
        lines = lines[:-1]
    if len(lines) >= 2 and _looks_like_trailing_artifact(lines[-1], lines[-2]):
        lines = lines[:-1]
    return "\n".join(lines).strip()


def _looks_like_trailing_artifact(candidate: str, previous: str) -> bool:
    if not candidate or not previous:
        return False
    if not re.fullmatch(r"[0-9]{1,2}", candidate):
        return False
    if re.search(r"(Version|Page)\s*$", previous, re.IGNORECASE):
        return False
    if re.search(r"[가-힣A-Za-z)]$", previous):
        return True
    return False


def _group_pymupdf_line_spans(spans: list[dict[str, Any]]) -> list[tuple[float, float, list[str]]]:
    groups: list[tuple[float, float, list[str]]] = []
    for span in spans:
        text = _clean_pdf_text(span.get("text", ""))
        if not text:
            continue
        x0, _y0, x1, _y1 = [float(v) for v in span.get("bbox", [0.0, 0.0, 0.0, 0.0])]
        if not groups:
            groups.append((x0, x1, [text]))
            continue
        prev_x0, prev_x1, prev_texts = groups[-1]
        gap = x0 - prev_x1
        if gap > max(float(span.get("size", 10.0)) * 1.8, 20.0):
            groups.append((x0, x1, [text]))
            continue
        prev_texts.append(text)
        groups[-1] = (prev_x0, x1, prev_texts)
    return groups


def _line_bbox(spans: list[dict[str, Any]]) -> tuple[float, float, float, float]:
    boxes = [[float(v) for v in span.get("bbox", [0.0, 0.0, 0.0, 0.0])] for span in spans if span.get("bbox")]
    if not boxes:
        return (0.0, 0.0, 0.0, 0.0)
    return (
        min(box[0] for box in boxes),
        min(box[1] for box in boxes),
        max(box[2] for box in boxes),
        max(box[3] for box in boxes),
    )


def _align_groups_to_columns(
    groups: list[tuple[float, float, list[str]]],
    column_signatures: list[float],
    *,
    allow_expand: bool,
) -> tuple[list[str] | None, float]:
    if not column_signatures:
        return None, 0.0

    aligned = [""] * len(column_signatures)
    matched = 0
    for x0, x1, texts in groups:
        center = (x0 + x1) / 2
        index = min(range(len(column_signatures)), key=lambda idx: abs(column_signatures[idx] - center))
        tolerance = max(35.0, (x1 - x0) * 1.5)
        if abs(column_signatures[index] - center) > tolerance:
            if not allow_expand or len(column_signatures) >= 6:
                return None, 0.0
            column_signatures.append(center)
            aligned.append("")
            index = len(column_signatures) - 1
        else:
            matched += 1
        content = " ".join(texts).strip()
        aligned[index] = f"{aligned[index]}\n{content}".strip() if aligned[index] else content

    if not any(cell for cell in aligned):
        return None, 0.0
    return aligned, matched / max(len(groups), 1)


def _should_merge_inferred_row(current: dict[str, Any], next_row: dict[str, Any]) -> bool:
    gap = next_row["bbox"][1] - current["bbox"][3]
    if gap > max(14.0, (current["bbox"][3] - current["bbox"][1]) * 0.85):
        return False

    current_filled = [index for index, cell in enumerate(current["cells"]) if cell.strip()]
    next_filled = [index for index, cell in enumerate(next_row["cells"]) if cell.strip()]
    if not current_filled or not next_filled:
        return False
    overlap = len(set(current_filled) & set(next_filled))
    return overlap >= max(1, min(len(current_filled), len(next_filled)) // 2)


def _merge_row_cells(current: list[str], next_row: list[str]) -> list[str]:
    width = max(len(current), len(next_row))
    merged: list[str] = []
    for index in range(width):
        left = current[index] if index < len(current) else ""
        right = next_row[index] if index < len(next_row) else ""
        if left and right:
            merged.append(f"{left}\n{right}")
        else:
            merged.append(left or right)
    return merged


def _merge_bboxes(
    left: tuple[float, float, float, float],
    right: tuple[float, float, float, float],
) -> tuple[float, float, float, float]:
    return (
        min(left[0], right[0]),
        min(left[1], right[1]),
        max(left[2], right[2]),
        max(left[3], right[3]),
    )


def _bbox_contains(outer: tuple[float, float, float, float], inner: tuple[float, float, float, float], tolerance: float = 6.0) -> bool:
    return (
        inner[0] >= outer[0] - tolerance
        and inner[1] >= outer[1] - tolerance
        and inner[2] <= outer[2] + tolerance
        and inner[3] <= outer[3] + tolerance
    )


def _bbox_area(bbox: tuple[float, float, float, float]) -> float:
    return max(0.0, bbox[2] - bbox[0]) * max(0.0, bbox[3] - bbox[1])


def _point_near_bbox(point: tuple[float, float], bbox: tuple[float, float, float, float], tolerance: float = 14.0) -> bool:
    x, y = point
    return (
        bbox[0] - tolerance <= x <= bbox[2] + tolerance
        and bbox[1] - tolerance <= y <= bbox[3] + tolerance
    )


def _points_close(left: tuple[float, float], right: tuple[float, float], tolerance: float = 12.0) -> bool:
    return abs(left[0] - right[0]) <= tolerance and abs(left[1] - right[1]) <= tolerance


def _segment_length(start: tuple[float, float], end: tuple[float, float]) -> float:
    return ((end[0] - start[0]) ** 2 + (end[1] - start[1]) ** 2) ** 0.5


def _box_label_text(box: dict[str, Any]) -> str:
    return str(box.get("label", "") or "").strip()


def _box_label_quality(box: dict[str, Any]) -> float:
    label = _box_label_text(box)
    if not label:
        return 0.0
    source = str(box.get("label_source", "") or "").strip().lower()
    if source == "native":
        return 1.0
    if source == "ocr":
        return 0.82
    return 0.65


def _find_connector_box_index(
    point: tuple[float, float],
    boxes: list[dict[str, Any]],
    *,
    tolerance: float = 14.0,
) -> int | None:
    candidates: list[tuple[int, int, float, float, float]] = []
    for index, box in enumerate(boxes):
        bbox = box.get("bbox", [])
        if not isinstance(bbox, list) or len(bbox) < 4:
            continue
        rect = tuple(float(value) for value in bbox[:4])
        if not _point_near_bbox(point, rect, tolerance=tolerance):
            continue
        label_quality = _box_label_quality(box)
        area = _bbox_area(rect)
        distance = min(
            abs(point[0] - rect[0]),
            abs(point[0] - rect[2]),
            abs(point[1] - rect[1]),
            abs(point[1] - rect[3]),
        )
        inside = int(rect[0] <= point[0] <= rect[2] and rect[1] <= point[1] <= rect[3])
        candidates.append((index, inside, label_quality, -distance, -area))

    if not candidates:
        return None

    labeled_candidates = [candidate for candidate in candidates if candidate[2] > 0.0]
    search_space = labeled_candidates or candidates
    best = max(search_space, key=lambda item: (item[1], item[2], item[3], item[4]))
    return best[0]


def _cluster_connector_points(
    connector_segments: list[tuple[tuple[float, float], tuple[float, float]]],
    *,
    tolerance: float = 12.0,
) -> tuple[list[tuple[float, float]], list[tuple[int, int]]]:
    clusters: list[dict[str, Any]] = []
    clustered_segments: list[tuple[int, int]] = []

    def assign(point: tuple[float, float]) -> int:
        for index, cluster in enumerate(clusters):
            if _points_close(point, cluster["center"], tolerance=tolerance):
                cluster["points"].append(point)
                xs = [item[0] for item in cluster["points"]]
                ys = [item[1] for item in cluster["points"]]
                cluster["center"] = (sum(xs) / len(xs), sum(ys) / len(ys))
                return index
        clusters.append({"points": [point], "center": point})
        return len(clusters) - 1

    for start, end in connector_segments:
        clustered_segments.append((assign(start), assign(end)))

    return [cluster["center"] for cluster in clusters], clustered_segments


def _infer_diagram_edges(
    boxes: list[dict[str, Any]],
    connector_segments: list[tuple[tuple[float, float], tuple[float, float]]],
    arrow_segments: list[tuple[tuple[float, float], tuple[float, float]]] | None = None,
) -> list[dict[str, Any]]:
    if len(boxes) < 2 or not connector_segments:
        return []

    nodes, clustered_segments = _cluster_connector_points(connector_segments)
    adjacency: dict[int, set[int]] = {index: set() for index in range(len(nodes))}
    node_segments: dict[int, list[int]] = {index: [] for index in range(len(nodes))}
    for seg_index, (left, right) in enumerate(clustered_segments):
        adjacency[left].add(right)
        adjacency[right].add(left)
        node_segments[left].append(seg_index)
        node_segments[right].append(seg_index)

    edges: list[dict[str, Any]] = []
    seen_edges: set[tuple[int, int, str]] = set()
    visited: set[int] = set()
    arrow_segments = arrow_segments or []
    segment_lengths = [_segment_length(nodes[left], nodes[right]) for left, right in clustered_segments]

    def is_routing_box(box_index: int, degree: int) -> bool:
        box = boxes[box_index]
        bbox = box.get("bbox", [])
        if not isinstance(bbox, list) or len(bbox) < 4:
            return False
        width = abs(float(bbox[2]) - float(bbox[0]))
        height = abs(float(bbox[3]) - float(bbox[1]))
        if _box_label_text(box):
            return False
        small = width <= 96.0 and height <= 36.0
        slender = max(width, height) >= 48.0 and min(width, height) <= 18.0
        junction_like = degree >= 3 or (degree >= 2 and (small or slender))
        return small or slender or junction_like

    def component_axis(payloads: list[dict[str, Any]]) -> str:
        xs = [nodes[item["node"]][0] for item in payloads]
        ys = [nodes[item["node"]][1] for item in payloads]
        if not xs or not ys:
            return "horizontal"
        return "horizontal" if (max(xs) - min(xs)) >= (max(ys) - min(ys)) else "vertical"

    def ordered_payloads(payloads: list[dict[str, Any]]) -> list[dict[str, Any]]:
        axis = component_axis(payloads)
        if axis == "vertical":
            return sorted(payloads, key=lambda item: (round(nodes[item["node"]][1], 2), round(nodes[item["node"]][0], 2)))
        return sorted(payloads, key=lambda item: (round(nodes[item["node"]][0], 2), round(nodes[item["node"]][1], 2)))

    def endpoint_arrow_score(endpoint: tuple[float, float], reference: tuple[float, float]) -> float:
        vx = reference[0] - endpoint[0]
        vy = reference[1] - endpoint[1]
        main_length = max((vx * vx + vy * vy) ** 0.5, 1.0)
        score = 0.0
        wing_vectors: list[tuple[float, float]] = []
        for start, end in arrow_segments:
            if _points_close(start, endpoint, tolerance=10.0):
                other = end
            elif _points_close(end, endpoint, tolerance=10.0):
                other = start
            else:
                continue
            length = _segment_length(endpoint, other)
            if length < 5.0 or length > 28.0:
                continue
            wx = other[0] - endpoint[0]
            wy = other[1] - endpoint[1]
            dot = (wx * vx + wy * vy) / (length * main_length)
            if dot <= 0.05:
                continue
            score += 1.0
            wing_vectors.append((wx / length, wy / length))
        if len(wing_vectors) >= 2:
            for left in range(len(wing_vectors) - 1):
                for right in range(left + 1, len(wing_vectors)):
                    cross = abs(
                        wing_vectors[left][0] * wing_vectors[right][1]
                        - wing_vectors[left][1] * wing_vectors[right][0]
                    )
                    if cross >= 0.25:
                        score += 0.35
            if len(wing_vectors) >= 3:
                score += 0.20
        return score

    def choose_root_index(payloads: list[dict[str, Any]], component_provenance: str) -> int:
        labeled_payloads = [item for item in payloads if _box_label_quality(boxes[item["box"]]) > 0.0]
        candidates = labeled_payloads or payloads
        if len(candidates) == 1:
            return candidates[0]["box"]
        axis = component_axis(candidates)

        def root_sort_key(item: dict[str, Any]) -> tuple[float, float, float, float]:
            x, y = nodes[item["node"]]
            primary = y if axis == "vertical" else x
            secondary = x if axis == "vertical" else y
            return (-_box_label_quality(boxes[item["box"]]), round(primary, 2), round(secondary, 2), _bbox_area(tuple(boxes[item["box"]].get("bbox", (0.0, 0.0, 0.0, 0.0))[:4])))

        if component_provenance == "branch":
            return sorted(candidates, key=root_sort_key)[0]["box"]
        return sorted(candidates, key=root_sort_key)[0]["box"]

    def edge_confidence(
        provenance: str,
        source_box: int,
        target_box: int,
        source_node: int,
        target_node: int,
        arrow_score: float,
        routing_nodes: int,
        component_length: float,
    ) -> float:
        base = {"direct": 0.80, "chain": 0.66, "branch": 0.62}[provenance]
        source_quality = _box_label_quality(boxes[source_box])
        target_quality = _box_label_quality(boxes[target_box])
        if source_quality and target_quality:
            base += 0.03 + (0.01 * min(source_quality, target_quality))
        elif source_quality or target_quality:
            base += 0.02
        else:
            base -= 0.04

        source_point = nodes[source_node]
        target_point = nodes[target_node]
        span = max(_segment_length(source_point, target_point), 1.0)
        straightness = min(span / max(component_length, span), 1.0)
        base += min(0.07, max(0.0, straightness - 0.55) * 0.14)

        if arrow_score > 0.0:
            base += min(0.12, 0.025 * arrow_score + (0.015 if arrow_score >= 2.0 else 0.0))
        if routing_nodes:
            base -= min(0.10, 0.02 * routing_nodes)
        if provenance == "branch" and source_quality and target_quality:
            base += 0.01
        return round(min(max(base, 0.55), 0.97), 2)

    for node_index in range(len(nodes)):
        if node_index in visited:
            continue
        stack = [node_index]
        component_nodes: list[int] = []
        component_segments: set[int] = set()
        while stack:
            current = stack.pop()
            if current in visited:
                continue
            visited.add(current)
            component_nodes.append(current)
            component_segments.update(node_segments.get(current, []))
            for neighbor in adjacency.get(current, set()):
                if neighbor not in visited:
                    stack.append(neighbor)

        endpoint_nodes = [index for index in component_nodes if len(adjacency.get(index, set())) <= 1]
        if len(endpoint_nodes) < 2:
            continue

        endpoint_payloads: list[dict[str, Any]] = []
        for endpoint in endpoint_nodes:
            box_index = _find_connector_box_index(nodes[endpoint], boxes)
            if box_index is not None:
                degree = len(adjacency.get(endpoint, set()))
                endpoint_payloads.append(
                    {
                        "node": endpoint,
                        "box": box_index,
                        "degree": degree,
                        "routing": is_routing_box(box_index, degree),
                    }
                )
        if len(endpoint_payloads) < 2:
            continue

        semantic_payloads = [item for item in endpoint_payloads if not item["routing"]]
        labeled_payloads = [item for item in semantic_payloads if _box_label_quality(boxes[item["box"]]) > 0.0]
        if len(labeled_payloads) >= 2:
            effective_payloads = labeled_payloads
        elif len(semantic_payloads) >= 2:
            effective_payloads = semantic_payloads
        else:
            effective_payloads = endpoint_payloads

        if len(component_segments) == 1:
            provenance = "direct"
        elif len(effective_payloads) > 2:
            provenance = "branch"
        else:
            provenance = "chain"

        ordered_effective_payloads = ordered_payloads(effective_payloads)
        payload_by_box = {item["box"]: item for item in ordered_effective_payloads}
        component_length = sum(segment_lengths[index] for index in component_segments) if component_segments else 1.0

        arrow_target_box: int | None = None
        arrow_score = 0.0
        for payload in ordered_effective_payloads:
            endpoint_node = payload["node"]
            others = [item for item in ordered_effective_payloads if item["node"] != endpoint_node]
            if not others:
                continue
            nearest_other = min(others, key=lambda item: _segment_length(nodes[endpoint_node], nodes[item["node"]]))
            score = endpoint_arrow_score(nodes[endpoint_node], nodes[nearest_other["node"]])
            if score > arrow_score:
                arrow_score = score
                arrow_target_box = payload["box"]

        routing_nodes = max(0, len(endpoint_payloads) - len(effective_payloads))

        if provenance == "branch":
            root_box = choose_root_index(effective_payloads, provenance)
            root_payload = payload_by_box.get(root_box)
            if arrow_target_box is not None and len(ordered_effective_payloads) >= 2 and arrow_target_box in payload_by_box:
                target_payload = payload_by_box[arrow_target_box]
                for payload in ordered_effective_payloads:
                    source_box = payload["box"]
                    if source_box == arrow_target_box:
                        continue
                    edge_key = (source_box, arrow_target_box, provenance)
                    if edge_key in seen_edges:
                        continue
                    seen_edges.add(edge_key)
                    confidence = edge_confidence(
                        provenance,
                        source_box,
                        arrow_target_box,
                        payload["node"],
                        target_payload["node"],
                        arrow_score,
                        routing_nodes,
                        component_length,
                    )
                    edges.append(
                        {
                            "from_index": source_box + 1,
                            "to_index": arrow_target_box + 1,
                            "from_label": str(boxes[source_box].get("label", "") or ""),
                            "to_label": str(boxes[arrow_target_box].get("label", "") or ""),
                            "provenance": provenance,
                            "direction_hint": "arrowhead",
                            "confidence": confidence,
                            "routing_nodes": routing_nodes,
                        }
                    )
            elif root_payload is not None:
                for payload in ordered_effective_payloads:
                    target_box = payload["box"]
                    if target_box == root_box:
                        continue
                    edge_key = (root_box, target_box, provenance)
                    if edge_key in seen_edges:
                        continue
                    seen_edges.add(edge_key)
                    confidence = edge_confidence(
                        provenance,
                        root_box,
                        target_box,
                        root_payload["node"],
                        payload["node"],
                        arrow_score,
                        routing_nodes,
                        component_length,
                    )
                    edges.append(
                        {
                            "from_index": root_box + 1,
                            "to_index": target_box + 1,
                            "from_label": str(boxes[root_box].get("label", "") or ""),
                            "to_label": str(boxes[target_box].get("label", "") or ""),
                            "provenance": provenance,
                            "direction_hint": "spatial-branch",
                            "confidence": confidence,
                            "routing_nodes": routing_nodes,
                        }
                    )
            continue

        first_payload = ordered_effective_payloads[0]
        last_payload = ordered_effective_payloads[-1]
        first_node, first_box = first_payload["node"], first_payload["box"]
        last_node, last_box = last_payload["node"], last_payload["box"]
        if first_box == last_box:
            continue
        first_point = nodes[first_node]
        last_point = nodes[last_node]
        if arrow_target_box is not None:
            if first_box == arrow_target_box:
                first_box, last_box = last_box, first_box
                first_node, last_node = last_node, first_node
            direction_hint = "arrowhead"
        elif provenance == "direct":
            direction_hint = "segment"
        else:
            direction_hint = "spatial"
            if abs(last_point[1] - first_point[1]) > abs(last_point[0] - first_point[0]):
                if first_point[1] > last_point[1]:
                    first_box, last_box = last_box, first_box
                    first_node, last_node = last_node, first_node
            elif first_point[0] > last_point[0]:
                first_box, last_box = last_box, first_box
                first_node, last_node = last_node, first_node

        first_label = str(boxes[first_box].get("label", "") or "")
        last_label = str(boxes[last_box].get("label", "") or "")
        confidence = edge_confidence(
            provenance,
            first_box,
            last_box,
            first_node,
            last_node,
            arrow_score,
            routing_nodes,
            component_length,
        )
        edge_key = (first_box, last_box, provenance)
        if edge_key in seen_edges:
            continue
        seen_edges.add(edge_key)
        edges.append(
            {
                "from_index": first_box + 1,
                "to_index": last_box + 1,
                "from_label": first_label,
                "to_label": last_label,
                "provenance": provenance,
                "direction_hint": direction_hint,
                "confidence": confidence,
                "routing_nodes": routing_nodes,
            }
        )

    return edges


def _table_columns_are_compatible(left: list[float], right: list[float]) -> bool:
    if not left or not right or abs(len(left) - len(right)) > 1:
        return False
    pairs = zip(left, right)
    drift = [abs(a - b) for a, b in pairs]
    return bool(drift) and sum(drift) / len(drift) <= 36.0


def _render_pymupdf_text_block(block: dict[str, Any]) -> tuple[str, str, str]:
    column_rows, confidence = _extract_block_column_rows(block)
    if len(column_rows) >= 2 and confidence >= 0.55:
        return _render_column_rows(column_rows)

    lines = []
    max_font = 0.0

    for line in block.get("lines", []):
        spans = line.get("spans", [])
        if not spans:
            continue
        line_text = _join_pymupdf_spans(spans)
        line_text = _clean_pdf_text(line_text)
        if not line_text or _is_pdf_noise(line_text):
            continue
        lines.append(line_text)
        max_font = max(max_font, max((float(span.get("size", 0.0)) for span in spans), default=0.0))

    if not lines:
        return "", "", ""

    if _looks_like_pdf_list_item(lines[0]):
        items = []
        markdown_items = []
        for line in lines:
            content = re.sub(r"^[\u2022\u25aa\u25cf▪\-]\s*", "", line).strip()
            items.append(f"<li>{escape(content)}</li>")
            markdown_items.append(f"- {content}")
        return "\n".join(lines), "<ul>\n" + "\n".join(items) + "\n</ul>", "\n".join(markdown_items)

    text = "\n".join(lines)
    if _looks_like_pdf_heading(lines[0]) or max_font >= 18:
        level = "h3" if max_font >= 22 else "h4"
        markdown_level = "###" if level == "h3" else "####"
        heading = " ".join(lines)
        return text, f"<{level}>{escape(heading)}</{level}>", f"{markdown_level} {heading}"

    paragraphs = _merge_pdf_paragraph_lines(lines)
    html = "\n".join(f"<p>{escape(paragraph)}</p>" for paragraph in paragraphs)
    markdown = "\n\n".join(paragraphs)
    return "\n".join(paragraphs), html, markdown


def _extract_block_column_rows(block: dict[str, Any]) -> tuple[list[list[str]], float]:
    rows: list[list[str]] = []
    column_signatures: list[float] = []
    aligned_rows = 0
    for line in block.get("lines", []):
        spans = [span for span in line.get("spans", []) if (span.get("text") or "").strip()]
        groups = _group_pymupdf_line_spans(spans)
        if len(groups) >= 2 and len(groups) <= 5:
            if not column_signatures:
                column_signatures = [(x0 + x1) / 2 for x0, x1, _texts in groups]
            aligned, match_ratio = _align_groups_to_columns(
                groups,
                column_signatures,
                allow_expand=len(column_signatures) < 5,
            )
            if aligned:
                rows.append(aligned)
                aligned_rows += match_ratio >= 0.55

    if len(rows) < 2:
        return [], 0.0

    widths = {len(row) for row in rows}
    if len(widths) != 1:
        return [], 0.0
    if list(widths)[0] < 2:
        return [], 0.0
    total_lines = max(len(block.get("lines", [])), 1)
    confidence = aligned_rows / total_lines
    return rows, confidence


def _align_row_to_columns(
    groups: list[tuple[float, float, list[str]]],
    column_signatures: list[float],
) -> list[str] | None:
    return _align_groups_to_columns(groups, column_signatures, allow_expand=len(column_signatures) < 5)[0]


def _render_column_rows(rows: list[list[str]]) -> tuple[str, str, str]:
    html, text, markdown = _render_table_rows(rows, None, inferred=False)
    return text, html, markdown


def _extract_pymupdf_diagram_primitives(page: Any, page_dict: dict[str, Any]) -> dict[str, Any]:
    try:
        drawings = page.get_drawings()
    except Exception:
        drawings = []

    page_width = float(page.rect.width or 0.0)
    page_height = float(page.rect.height or 0.0)
    text_blocks: list[tuple[tuple[float, float, float, float], str]] = []
    for block in page_dict.get("blocks", []):
        if block.get("type") != 0:
            continue
        bbox = tuple(float(value) for value in block.get("bbox", (0.0, 0.0, 0.0, 0.0)))
        block_width = max(0.0, bbox[2] - bbox[0])
        block_height = max(0.0, bbox[3] - bbox[1])
        if page_width > 0 and page_height > 0:
            is_edge_band = bbox[1] <= page_height * 0.12 or bbox[3] >= page_height * 0.88
            is_wide_strip = block_width >= page_width * 0.6 and block_height <= page_height * 0.12
            if is_edge_band and is_wide_strip:
                continue
        text, _html, _markdown = _render_pymupdf_text_block(block)
        if text:
            text_blocks.append((bbox, text.replace("\n", " ").strip()))

    rects: list[tuple[float, float, float, float]] = []
    connectors = 0
    line_segments: list[tuple[tuple[float, float], tuple[float, float]]] = []
    connector_segments: list[tuple[tuple[float, float], tuple[float, float]]] = []
    page_bbox = (0.0, 0.0, page_width, page_height)
    page_area = max(_bbox_area(page_bbox), 1.0)
    for drawing in drawings:
        for item in drawing.get("items", []):
            operator = item[0] if item else ""
            if operator == "re" and len(item) >= 2:
                rect = item[1]
                if hasattr(rect, "x0"):
                    rects.append((float(rect.x0), float(rect.y0), float(rect.x1), float(rect.y1)))
            elif operator == "l" and len(item) >= 3:
                p1, p2 = item[1], item[2]
                start = (float(getattr(p1, "x", 0.0)), float(getattr(p1, "y", 0.0)))
                end = (float(getattr(p2, "x", 0.0)), float(getattr(p2, "y", 0.0)))
                dx = abs(start[0] - end[0])
                dy = abs(start[1] - end[1])
                if dx >= 8.0 or dy >= 8.0:
                    line_segments.append((start, end))
                if dx >= 24.0 or dy >= 24.0:
                    connectors += 1
                    connector_segments.append((start, end))

    boxes: list[dict[str, Any]] = []
    used_labels: set[str] = set()
    for rect in rects:
        rect_width = max(0.0, rect[2] - rect[0])
        rect_height = max(0.0, rect[3] - rect[1])
        if _bbox_area(rect) / page_area >= 0.72:
            continue
        if page_width > 0 and page_height > 0:
            is_edge_band = rect[1] <= page_height * 0.12 or rect[3] >= page_height * 0.88
            is_wide_strip = rect_width >= page_width * 0.6 and rect_height <= page_height * 0.12
            if is_edge_band and is_wide_strip:
                continue
        labels = []
        contained_blocks = 0
        for text_bbox, text in text_blocks:
            if text and _bbox_contains(rect, text_bbox):
                contained_blocks += 1
                labels.append(text)
        if contained_blocks >= 4:
            continue
        label = " | ".join(labels).strip()
        if not label and page_width > 0 and page_height > 0:
            is_small_unlabeled = rect_width <= page_width * 0.22 and rect_height <= page_height * 0.06
            connected = any(
                _point_near_bbox(start, rect) or _point_near_bbox(end, rect)
                for start, end in connector_segments
            )
            if is_small_unlabeled and not connected:
                continue
        dedupe_key = f"{round(rect[0],1)}:{round(rect[1],1)}:{label}"
        if dedupe_key in used_labels:
            continue
        used_labels.add(dedupe_key)
        boxes.append(
            {
                "bbox": [round(value, 2) for value in rect],
                "label": label,
                "label_source": "native" if label else "",
            }
        )

    unlabeled = sum(1 for box in boxes if not box["label"])
    arrow_segments = [segment for segment in line_segments if _segment_length(segment[0], segment[1]) <= 28.0]
    edges = _infer_diagram_edges(boxes, connector_segments, arrow_segments)
    return {
        "boxes": boxes[:12],
        "connectors": connectors,
        "connector_segments": [
            [[round(start[0], 2), round(start[1], 2)], [round(end[0], 2), round(end[1], 2)]]
            for start, end in connector_segments[:32]
        ],
        "edges": edges[:24],
        "unlabeled_boxes": unlabeled,
    }


def _merge_ocr_labels_into_diagram(diagram: dict[str, Any], ocr_payload: dict[str, Any], page: Any) -> dict[str, Any]:
    boxes = [dict(box) for box in diagram.get("boxes", [])]
    ocr_words = ocr_payload.get("words", [])
    image_width = int(ocr_payload.get("image_width", 0) or 0)
    image_height = int(ocr_payload.get("image_height", 0) or 0)
    if not boxes or not isinstance(ocr_words, list) or image_width <= 0 or image_height <= 0:
        return diagram

    page_width = float(page.rect.width or 1.0)
    page_height = float(page.rect.height or 1.0)
    scale_x = page_width / image_width
    scale_y = page_height / image_height
    label_retry = bool(isinstance(ocr_payload.get("strategy"), dict) and ocr_payload["strategy"].get("label_retry"))
    strict_floor = 45.0 if label_retry else 55.0
    fallback_floor = 18.0 if label_retry else 25.0
    fallback_limit = 32 if label_retry else 24

    for box in boxes:
        if box.get("label"):
            continue
        bbox = box.get("bbox", [0.0, 0.0, 0.0, 0.0])
        if not isinstance(bbox, list) or len(bbox) < 4:
            continue
        rect = tuple(float(value) for value in bbox[:4])
        strict_labels: list[tuple[float, float, float, str]] = []
        fallback_labels: list[tuple[float, float, float, str]] = []
        for word in ocr_words:
            if not isinstance(word, dict):
                continue
            text = str(word.get("text", "")).strip()
            confidence = float(word.get("confidence", 0.0) or 0.0)
            wbbox = word.get("bbox", [])
            if not text or not isinstance(wbbox, list) or len(wbbox) < 4:
                continue
            word_bbox = (
                float(wbbox[0]) * scale_x,
                float(wbbox[1]) * scale_y,
                float(wbbox[2]) * scale_x,
                float(wbbox[3]) * scale_y,
            )
            if not _bbox_contains(rect, word_bbox, tolerance=8.0):
                continue
            word_order = (word_bbox[1], word_bbox[0])
            if confidence >= strict_floor:
                strict_labels.append((word_order[0], word_order[1], confidence, text))
            elif confidence >= fallback_floor and len(text) <= fallback_limit and _ocr_word_is_label_like(
                text,
                confidence,
                layout_hint="diagram" if label_retry else "text",
                label_focus=label_retry,
            ):
                fallback_labels.append((word_order[0], word_order[1], confidence, text))
        strict_labels.sort(key=lambda item: (round(item[0], 1), round(item[1], 1)))
        labels = [text for _y, _x, _confidence, text in strict_labels]
        if not labels and fallback_labels:
            fallback_labels.sort(key=lambda item: (round(item[0], 1), round(item[1], 1), -item[2]))
            labels = [text for _y, _x, _confidence, text in fallback_labels[:4]]
        if labels:
            box["label"] = " ".join(labels).strip()
            box["label_source"] = "ocr"

    merged = {
        "boxes": boxes,
        "connectors": int(diagram.get("connectors", 0) or 0),
        "connector_segments": list(diagram.get("connector_segments", [])),
        "edges": list(diagram.get("edges", [])),
        "unlabeled_boxes": sum(1 for box in boxes if not box.get("label")),
    }
    return merged


def _render_diagram_summary_html(diagram: dict[str, Any]) -> str:
    boxes = diagram.get("boxes", [])
    connectors = int(diagram.get("connectors", 0) or 0)
    edges = diagram.get("edges", [])
    unlabeled = int(diagram.get("unlabeled_boxes", 0) or 0)
    if not boxes and not edges:
        return ""
    items = []
    for index, box in enumerate(boxes[:8], start=1):
        label = box.get("label") or "(unlabeled)"
        source = f" [{box.get('label_source')}]" if box.get("label_source") else ""
        items.append(f"<li>Box {index}{escape(source)}: {escape(str(label))}</li>")
    edge_items = []
    for edge in edges[:6]:
        source_label = escape(str(edge.get("from_label") or f"Box {edge.get('from_index', '?')}"))
        target_label = escape(str(edge.get("to_label") or f"Box {edge.get('to_index', '?')}"))
        fragments = []
        if edge.get("provenance"):
            fragments.append(str(edge["provenance"]))
        if edge.get("confidence") is not None:
            fragments.append(f"conf {edge['confidence']}")
        note = f" <small>({' / '.join(fragments)})</small>" if fragments else ""
        edge_items.append(f"<li>{source_label} -&gt; {target_label}{note}</li>")
    meta = (
        f"<p><small>Diagram hints | boxes: {len(boxes)} | connectors: {connectors} "
        f"| edges: {len(edges)} | unlabeled boxes: {unlabeled}</small></p>"
    )
    if not items:
        return f"<div class=\"pdf-diagram\">{meta}</div>"
    html = f"<div class=\"pdf-diagram\">{meta}<ul>{''.join(items)}</ul>"
    if edge_items:
        html += f"<p><small>Flow</small></p><ul>{''.join(edge_items)}</ul>"
    html += "</div>"
    return html


def _render_diagram_summary_markdown(diagram: dict[str, Any]) -> str:
    boxes = diagram.get("boxes", [])
    connectors = int(diagram.get("connectors", 0) or 0)
    edges = diagram.get("edges", [])
    unlabeled = int(diagram.get("unlabeled_boxes", 0) or 0)
    if not boxes and not edges:
        return ""
    parts = [
        f"> Diagram hints | boxes: {len(boxes)} | connectors: {connectors} | edges: {len(edges)} | unlabeled boxes: {unlabeled}"
    ]
    for index, box in enumerate(boxes[:6], start=1):
        label = box.get("label") or "(unlabeled)"
        source = f" [{box.get('label_source')}]" if box.get("label_source") else ""
        parts.append(f"- Box {index}{source}: {label}")
    for edge in edges[:6]:
        source_label = edge.get("from_label") or f"Box {edge.get('from_index', '?')}"
        target_label = edge.get("to_label") or f"Box {edge.get('to_index', '?')}"
        fragments = []
        if edge.get("provenance"):
            fragments.append(str(edge["provenance"]))
        if edge.get("confidence") is not None:
            fragments.append(f"conf {edge['confidence']}")
        suffix = f" ({', '.join(fragments)})" if fragments else ""
        parts.append(f"- Flow: {source_label} -> {target_label}{suffix}")
    return "\n".join(parts)


def _block_bbox_area(block: dict[str, Any]) -> float:
    bbox = block.get("bbox", ())
    if not isinstance(bbox, (list, tuple)) or len(bbox) < 4:
        return 0.0
    x0, y0, x1, y1 = [float(value) for value in bbox[:4]]
    return max(0.0, x1 - x0) * max(0.0, y1 - y0)


def _count_block_lines(block: dict[str, Any]) -> int:
    if block.get("type") != 0:
        return 0
    return sum(1 for line in block.get("lines", []) if any((span.get("text") or "").strip() for span in line.get("spans", [])))


def _count_block_text_chars(block: dict[str, Any]) -> int:
    if block.get("type") != 0:
        return 0
    count = 0
    for line in block.get("lines", []):
        for span in line.get("spans", []):
            count += len((span.get("text") or "").strip())
    return count


def _analyze_pdf_page(blocks: list[dict[str, Any]], table_count: int) -> PageAnalysis:
    text_blocks = sum(1 for block in blocks if block.get("type") == 0)
    image_blocks = sum(1 for block in blocks if block.get("type") == 1)
    drawing_blocks = sum(1 for block in blocks if block.get("type") not in {0, 1})
    text_line_count = sum(_count_block_lines(block) for block in blocks)
    text_char_count = sum(_count_block_text_chars(block) for block in blocks)

    text_area = sum(_block_bbox_area(block) for block in blocks if block.get("type") == 0)
    image_area = sum(_block_bbox_area(block) for block in blocks if block.get("type") == 1)
    drawing_area = sum(_block_bbox_area(block) for block in blocks if block.get("type") not in {0, 1})
    total_area = max(text_area + image_area + drawing_area, 1.0)
    text_area_ratio = text_area / total_area
    image_area_ratio = image_area / total_area
    drawing_area_ratio = drawing_area / total_area

    text_score = (
        min(text_blocks / 6, 1.0) * 0.28
        + min(text_line_count / 18, 1.0) * 0.28
        + min(text_char_count / 900, 1.0) * 0.24
        + min(text_area_ratio, 1.0) * 0.2
    )
    table_presence = min(table_count / 2, 1.0)
    table_score = table_presence * (
        0.58
        + min(text_blocks / 6, 1.0) * 0.12
        + min(text_line_count / 14, 1.0) * 0.1
        + min(1.0 - image_area_ratio, 1.0) * 0.08
        + min(1.0 - drawing_area_ratio, 1.0) * 0.12
    )
    diagram_score = (
        min(image_blocks / 2, 1.0) * 0.22
        + min(drawing_blocks / 4, 1.0) * 0.22
        + min(image_area_ratio, 1.0) * 0.28
        + min(drawing_area_ratio, 1.0) * 0.2
        + max(0.0, 1.0 - min(text_line_count / 16, 1.0)) * 0.08
    )

    signal_scores = {
        "text": round(min(text_score, 0.99), 2),
        "table": round(min(table_score, 0.99), 2),
        "diagram": round(min(diagram_score, 0.99), 2),
    }
    ranked = sorted(signal_scores.items(), key=lambda item: item[1], reverse=True)
    dominant_signal, dominant_score = ranked[0]
    secondary_signal, secondary_score = ranked[1]

    mixed_candidate = (
        dominant_score >= 0.48
        and secondary_score >= 0.38
        and dominant_signal != secondary_signal
        and abs(dominant_score - secondary_score) <= 0.14
    )
    if not mixed_candidate:
        mixed_candidate = (
            signal_scores["text"] >= 0.44
            and signal_scores["diagram"] >= 0.3
            and (image_blocks + drawing_blocks) >= 1
            and text_blocks >= 1
            and abs(signal_scores["text"] - signal_scores["diagram"]) <= 0.22
        )

    if mixed_candidate:
        layout = "mixed"
        mixed_signals = sorted(signal_scores.items(), key=lambda item: item[1], reverse=True)[:2]
        confidence = min(0.56 + sum(score for _name, score in mixed_signals) / len(mixed_signals) * 0.32, 0.9)
        dominant_signal = "+".join(name for name, _score in mixed_signals)
    else:
        layout = dominant_signal
        confidence = min(0.5 + dominant_score * 0.42, 0.97)

    return PageAnalysis(
        layout=layout,
        layout_confidence=round(confidence, 2),
        text_blocks=text_blocks,
        image_blocks=image_blocks,
        drawing_blocks=drawing_blocks,
        table_count=table_count,
        text_line_count=text_line_count,
        text_char_count=text_char_count,
        dominant_signal=dominant_signal,
        signal_scores=signal_scores,
    )


def _join_pymupdf_spans(spans: list[dict[str, Any]]) -> str:
    pieces: list[str] = []
    last_right: float | None = None
    for span in spans:
        text = span.get("text", "")
        if not text:
            continue
        left = float(span.get("bbox", [0, 0, 0, 0])[0])
        right = float(span.get("bbox", [0, 0, 0, 0])[2])
        if pieces and last_right is not None and left - last_right > max(5.0, float(span.get("size", 10.0)) * 0.35):
            pieces.append(" ")
        pieces.append(text)
        last_right = right
    return "".join(pieces)


def _merge_pdf_paragraph_lines(lines: list[str]) -> list[str]:
    paragraphs: list[str] = []
    current = lines[0]
    for line in lines[1:]:
        if _starts_new_pdf_block(current, line):
            paragraphs.append(current)
            current = line
            continue
        if _should_merge_pdf_lines(current, line):
            joiner = "" if _needs_tight_join(current, line) else " "
            current = f"{current}{joiner}{line}".strip()
        else:
            paragraphs.append(current)
            current = line
    paragraphs.append(current)
    return paragraphs


def _extract_pdf_page_ocr_payload(
    page: Any,
    source_path: Path,
    document: Any,
    page_index: int,
    *,
    cache_dir: Path | None = None,
    cache_report: dict[str, dict[str, Any]] | None = None,
    force: bool = False,
    dpi: int = 200,
    layout_hint: str = "text",
) -> dict[str, str]:
    variant = "force" if force else "auto"
    profile = _build_ocr_profile(layout_hint, force=force)
    profile_token = _ocr_profile_cache_token(profile)
    cache_key = _build_page_cache_key("ocr", source_path, page_index, dpi=dpi, variant=variant, profile=profile_token)
    cache_meta = _build_page_cache_metadata(source_path, document, page_index, dpi=dpi, variant=variant, profile=profile_token)
    cached = _load_ocr_cache(cache_dir, cache_key, cache_meta, cache_report)
    if cached:
        cached["status"] = "cache-hit"
        return cached
    page_dict = page.get_text("dict", sort=True)
    diagram = _extract_pymupdf_diagram_primitives(page, page_dict)
    cache_path = _cache_file_path(cache_dir, "ocr", cache_key)
    with _cache_lock(cache_path):
        if cache_path and cache_path.exists():
            cached = _load_ocr_cache(cache_dir, cache_key, cache_meta, cache_report)
            if cached:
                cached["status"] = "cache-hit"
                return cached
        raster = _get_cached_page_raster(page, source_path, document, page_index, dpi, cache_dir, cache_report)
        if not raster:
            return {"text": "", "overlay": "", "status": "ocr-failed", "confidence_summary": {}}
        payload = _extract_ocr_from_raster(
            (
                0,
                base64.b64decode(raster["data"]),
                force,
                dpi,
                profile["key"],
                diagram,
                (float(page.rect.width or 0.0), float(page.rect.height or 0.0)),
            )
        )[1]
        payload["status"] = "ocr-force" if force else "ocr-auto"
        _save_ocr_cache(cache_dir, cache_key, cache_meta, payload, cache_report)
        return payload


def _prepare_cache_dir(cache_dir: Path | None) -> Path | None:
    if not cache_dir:
        return None
    cache_dir.mkdir(parents=True, exist_ok=True)
    return cache_dir


def _document_cache_signature(source_path: Path, document: Any) -> str:
    meta = getattr(document, "metadata", {}) or {}
    try:
        stat = source_path.stat()
        signature = f"{source_path.resolve()}|{stat.st_size}|{stat.st_mtime_ns}"
    except OSError:
        signature = str(source_path.resolve())
    fingerprint = "|".join(
        [
            signature,
            str(getattr(document, "page_count", "")),
            str(meta.get("format", "")),
            str(meta.get("title", "")),
            str(meta.get("author", "")),
        ]
    )
    return hashlib.sha256(fingerprint.encode("utf-8")).hexdigest()


def _build_page_cache_key(kind: str, source_path: Path, page_index: int, **extra: Any) -> str:
    parts = [kind, str(source_path.resolve()), str(page_index)]
    for key in sorted(extra):
        parts.append(f"{key}={extra[key]}")
    return hashlib.sha256("|".join(parts).encode("utf-8")).hexdigest()


def _build_page_cache_metadata(
    source_path: Path,
    document: Any,
    page_index: int,
    **extra: Any,
) -> dict[str, Any]:
    metadata = {
        "schema": CACHE_SCHEMA_VERSION,
        "signature": _document_cache_signature(source_path, document),
        "source": str(source_path.resolve()),
        "page_index": page_index,
    }
    metadata.update(extra)
    return metadata


def _cache_file_path(cache_dir: Path | None, kind: str, cache_key: str) -> Path | None:
    if not cache_dir:
        return None
    return _prepare_cache_dir(cache_dir / kind) / f"{cache_key}.json"


def _create_cache_report() -> dict[str, dict[str, Any]]:
    return {kind: {"hit": 0, "miss": 0, "stale": 0, "write": 0, "reasons": {}} for kind in ("ocr", "raster", "tables")}


def _record_cache_event(
    cache_report: dict[str, dict[str, Any]] | None,
    kind: str,
    event: str,
    reason: str | None = None,
) -> None:
    if cache_report is None:
        return
    bucket = cache_report.setdefault(kind, {"hit": 0, "miss": 0, "stale": 0, "write": 0, "reasons": {}})
    bucket[event] = int(bucket.get(event, 0)) + 1
    if reason:
        reasons = bucket.setdefault("reasons", {})
        reasons[reason] = int(reasons.get(reason, 0)) + 1


def _finalize_cache_report(cache_report: dict[str, dict[str, Any]]) -> dict[str, dict[str, Any]]:
    finalized: dict[str, dict[str, Any]] = {}
    for kind, stats in cache_report.items():
        entry = {
            "hit": int(stats.get("hit", 0)),
            "miss": int(stats.get("miss", 0)),
            "stale": int(stats.get("stale", 0)),
            "write": int(stats.get("write", 0)),
        }
        reasons = stats.get("reasons", {})
        if isinstance(reasons, dict) and reasons:
            entry["reasons"] = {str(key): int(value) for key, value in reasons.items()}
        finalized[kind] = entry
    return finalized


def _metadata_matches(expected: dict[str, Any], actual: dict[str, Any]) -> tuple[bool, str | None]:
    if int(actual.get("schema", -1)) != int(expected.get("schema", CACHE_SCHEMA_VERSION)):
        return False, "schema"
    if str(actual.get("signature", "")) != str(expected.get("signature", "")):
        return False, "signature"
    for key, value in expected.items():
        if key in {"schema", "signature"}:
            continue
        if actual.get(key) != value:
            return False, key
    return True, None


def _load_json_cache(
    cache_dir: Path | None,
    kind: str,
    cache_key: str,
    expected_metadata: dict[str, Any],
    cache_report: dict[str, dict[str, Any]] | None = None,
) -> Any | None:
    path = _cache_file_path(cache_dir, kind, cache_key)
    if not path or not path.exists():
        _record_cache_event(cache_report, kind, "miss", "missing")
        return None
    return _load_json_cache_from_path(path, kind, expected_metadata, cache_report)


def _load_json_cache_from_path(
    path: Path,
    kind: str,
    expected_metadata: dict[str, Any],
    cache_report: dict[str, dict[str, Any]] | None = None,
) -> Any | None:
    try:
        envelope = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        _record_cache_event(cache_report, kind, "stale", "unreadable")
        return None
    if not isinstance(envelope, dict):
        _record_cache_event(cache_report, kind, "stale", "invalid-envelope")
        return None
    metadata = envelope.get("metadata", {})
    payload = envelope.get("payload")
    if not isinstance(metadata, dict):
        _record_cache_event(cache_report, kind, "stale", "invalid-metadata")
        return None
    matches, reason = _metadata_matches(expected_metadata, metadata)
    if not matches:
        _record_cache_event(cache_report, kind, "stale", reason)
        return None
    _record_cache_event(cache_report, kind, "hit")
    return payload


def _save_json_cache(
    cache_dir: Path | None,
    kind: str,
    cache_key: str,
    metadata: dict[str, Any],
    payload: Any,
    cache_report: dict[str, dict[str, Any]] | None = None,
) -> None:
    path = _cache_file_path(cache_dir, kind, cache_key)
    if not path:
        return
    try:
        _write_json_atomic(path, {"metadata": metadata, "payload": payload})
        _record_cache_event(cache_report, kind, "write")
    except Exception:
        return


@contextmanager
def _cache_lock(path: Path | None):
    if not path:
        yield
        return
    lock_path = path.with_suffix(path.suffix + ".lock")
    lock_path.parent.mkdir(parents=True, exist_ok=True)
    handle = lock_path.open("a+", encoding="utf-8")
    try:
        try:
            import fcntl

            fcntl.flock(handle.fileno(), fcntl.LOCK_EX)
        except Exception:
            pass
        yield
    finally:
        try:
            import fcntl

            fcntl.flock(handle.fileno(), fcntl.LOCK_UN)
        except Exception:
            pass
        handle.close()


def _write_json_atomic(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(
        mode="w",
        encoding="utf-8",
        dir=path.parent,
        prefix=f".{path.name}.",
        suffix=".tmp",
        delete=False,
    ) as handle:
        temp_path = Path(handle.name)
        json.dump(payload, handle, ensure_ascii=False)
    os.replace(temp_path, path)


def _load_ocr_cache(
    cache_dir: Path | None,
    cache_key: str,
    expected_metadata: dict[str, Any],
    cache_report: dict[str, dict[str, Any]] | None = None,
) -> dict[str, str] | None:
    payload = _load_json_cache(cache_dir, "ocr", cache_key, expected_metadata, cache_report)
    if not isinstance(payload, dict):
        return None
    summary = payload.get("confidence_summary", {})
    if not isinstance(summary, dict):
        summary = {}
    return {
        "text": str(payload.get("text", "")),
        "overlay": str(payload.get("overlay", "")),
        "profile": str(payload.get("profile", "")),
        "strategy": payload.get("strategy", {}),
        "words": payload.get("words", []),
        "image_width": int(payload.get("image_width", 0) or 0),
        "image_height": int(payload.get("image_height", 0) or 0),
        "confidence_summary": summary,
    }


def _save_ocr_cache(
    cache_dir: Path | None,
    cache_key: str,
    metadata: dict[str, Any],
    payload: dict[str, str],
    cache_report: dict[str, dict[str, Any]] | None = None,
) -> None:
    _save_json_cache(cache_dir, "ocr", cache_key, metadata, payload, cache_report)


def _build_ocr_profile(layout_hint: str, *, force: bool) -> dict[str, Any]:
    normalized = layout_hint if layout_hint in {"text", "table", "diagram", "mixed"} else "text"
    if normalized == "table":
        return {
            "key": f"{'force' if force else 'auto'}-table",
            "psm_candidates": [6, 4, 11],
            "threshold": 182,
            "grayscale": True,
            "variant_policy": "focused-table",
            "word_confidence_floor": 40,
        }
    if normalized == "diagram":
        return {
            "key": f"{'force' if force else 'auto'}-diagram",
            "psm_candidates": [11, 6, 7],
            "threshold": 166,
            "grayscale": True,
            "variant_policy": "diagram-label",
            "word_confidence_floor": 22,
            "label_retry": True,
        }
    if normalized == "mixed":
        return {
            "key": f"{'force' if force else 'auto'}-mixed",
            "psm_candidates": [6, 11, 7],
            "threshold": 172,
            "grayscale": True,
            "variant_policy": "diagram-label",
            "word_confidence_floor": 24,
            "label_retry": True,
        }
    return {
        "key": f"{'force' if force else 'auto'}-text",
        "psm_candidates": [6, 3],
        "threshold": 188,
        "grayscale": True,
        "variant_policy": "fast-text",
        "word_confidence_floor": 40,
    }


def _build_ocr_profile_variants(profile: dict[str, Any]) -> list[dict[str, Any]]:
    base_threshold = int(profile.get("threshold", 0) or 0)
    policy = str(profile.get("variant_policy", "balanced"))
    if policy == "fast-text":
        variants = [
            {**profile, "variant": "base"},
            {**profile, "variant": "soft", "threshold": max(base_threshold - 18, 120)},
        ]
    elif policy == "focused-table":
        variants = [
            {**profile, "variant": "base"},
            {**profile, "variant": "strong", "threshold": min(base_threshold + 12, 235)},
        ]
    elif policy == "diagram-label":
        variants = [
            {**profile, "variant": "base"},
            {**profile, "variant": "soft", "threshold": max(base_threshold - 22, 110)},
            {
                **profile,
                "variant": "label",
                "threshold": max(base_threshold - 34, 96),
                "word_confidence_floor": min(int(profile.get("word_confidence_floor", 40) or 40), 18),
            },
        ]
    else:
        variants = [
            {**profile, "variant": "base"},
            {**profile, "variant": "soft", "threshold": max(base_threshold - 18, 120)},
            {**profile, "variant": "strong", "threshold": min(base_threshold + 12, 235)},
        ]
    deduped: list[dict[str, Any]] = []
    seen: set[tuple[int, str]] = set()
    for item in variants:
        key = (int(item.get("threshold", 0) or 0), str(item.get("variant", "")))
        if key in seen:
            continue
        seen.add(key)
        deduped.append(item)
    return deduped


def _should_stop_ocr_trials(summary: dict[str, Any], score: float) -> str | None:
    avg = float(summary.get("avg", 0.0) or 0.0)
    low_ratio = float(summary.get("low_confidence_ratio", 1.0) or 0.0)
    words = int(summary.get("words", 0) or 0)
    if avg >= 96.0 and low_ratio <= 0.015 and words >= 2:
        return "excellent-confidence"
    if avg >= 93.0 and low_ratio <= 0.0 and words >= 5 and score >= 92.5:
        return "high-confidence-complete"
    return None


def _should_expand_ocr_variants(
    profile: dict[str, Any],
    tried_variants: list[str],
    best_summary: dict[str, Any] | None,
) -> list[dict[str, Any]]:
    if best_summary is None:
        return []
    policy = str(profile.get("variant_policy", "balanced"))
    tried = set(tried_variants)
    avg = float(best_summary.get("avg", 0.0) or 0.0)
    low_ratio = float(best_summary.get("low_confidence_ratio", 1.0) or 0.0)
    words = int(best_summary.get("words", 0) or 0)
    threshold = int(profile.get("threshold", 0) or 0)
    if policy == "fast-text" and "strong" not in tried and (avg < 78.0 or low_ratio >= 0.18 or words <= 2):
        return [{**profile, "variant": "strong", "threshold": min(threshold + 12, 235)}]
    if policy == "focused-table" and "soft" not in tried and (avg < 80.0 or low_ratio >= 0.14):
        return [{**profile, "variant": "soft", "threshold": max(threshold - 18, 120)}]
    if policy == "diagram-label":
        if "label" not in tried and (avg < 86.0 or low_ratio >= 0.1 or words <= 6):
            return [
                {
                    **profile,
                    "variant": "label",
                    "threshold": max(threshold - 30, 96),
                    "word_confidence_floor": min(int(profile.get("word_confidence_floor", 40) or 40), 18),
                }
            ]
        if "soft" not in tried and avg < 79.0 and low_ratio >= 0.06:
            return [{**profile, "variant": "soft", "threshold": max(threshold - 22, 110)}]
    return []


def _normalize_ocr_word_text(text: str) -> str:
    cleaned = re.sub(r"\s+", " ", text or "").strip()
    cleaned = cleaned.strip("·•:;,_-–—/\\|[]{}()<>")
    return cleaned.strip()


def _ocr_word_is_label_like(text: str, confidence: float, *, layout_hint: str, label_focus: bool) -> bool:
    compact = re.sub(r"\s+", "", text)
    if not compact:
        return False
    if not re.search(r"[0-9A-Za-z\uac00-\ud7a3]", compact):
        return False
    if re.fullmatch(r"[_\W]+", compact):
        return False
    if len(compact) <= 1:
        return confidence >= (18.0 if label_focus or layout_hint in {"diagram", "mixed"} else 28.0)
    if len(compact) <= 3:
        return confidence >= (14.0 if label_focus or layout_hint in {"diagram", "mixed"} else 24.0)
    if len(compact) <= 6 and layout_hint in {"diagram", "mixed"}:
        return confidence >= 12.0 if label_focus else confidence >= 20.0
    return True


def _extract_ocr_words(
    ocr_data: dict[str, list[Any]],
    *,
    min_confidence: float = 40.0,
    layout_hint: str = "text",
    label_focus: bool = False,
) -> list[dict[str, Any]]:
    words: list[dict[str, Any]] = []
    total = len(ocr_data.get("text", []))
    effective_floor = float(min_confidence)
    if layout_hint in {"diagram", "mixed"}:
        effective_floor = min(effective_floor, 28.0)
    if label_focus:
        effective_floor = min(effective_floor, 18.0)
    for index in range(total):
        text = _normalize_ocr_word_text(str(ocr_data.get("text", [""] * total)[index] or ""))
        confidence = ocr_data.get("conf", ["-1"] * total)[index]
        try:
            score = float(confidence)
        except (TypeError, ValueError):
            score = -1.0
        if not text:
            continue
        if score < effective_floor and not _ocr_word_is_label_like(text, score, layout_hint=layout_hint, label_focus=label_focus):
            continue
        left = int(ocr_data.get("left", [0] * total)[index] or 0)
        top = int(ocr_data.get("top", [0] * total)[index] or 0)
        width = int(ocr_data.get("width", [0] * total)[index] or 0)
        height = int(ocr_data.get("height", [0] * total)[index] or 0)
        if width <= 0 or height <= 0:
            continue
        if len(re.sub(r"\s+", "", text)) <= 1 and score < 12.0 and not label_focus:
            continue
        words.append(
            {
                "text": text,
                "confidence": round(score, 2),
                "bbox": [left, top, left + width, top + height],
            }
        )
    return words


def _build_structured_ocr_text(
    words: list[dict[str, Any]],
    *,
    image_width: int,
    image_height: int,
    layout_hint: str,
) -> str:
    if not words or image_width <= 0 or image_height <= 0:
        return ""

    prepared: list[dict[str, Any]] = []
    heights: list[float] = []
    for word in words:
        if not isinstance(word, dict):
            continue
        text = str(word.get("text", "")).strip()
        bbox = word.get("bbox", [])
        if not text or not isinstance(bbox, list) or len(bbox) != 4:
            continue
        x0, y0, x1, y1 = [float(value) for value in bbox]
        if x1 <= x0 or y1 <= y0:
            continue
        heights.append(y1 - y0)
        prepared.append({"text": text, "bbox": [x0, y0, x1, y1]})
    if not prepared:
        return ""

    median_height = sorted(heights)[len(heights) // 2] if heights else 12.0
    line_tolerance = max(8.0, median_height * (0.65 if layout_hint in {"diagram", "mixed"} else 0.8))
    prepared.sort(key=lambda item: ((item["bbox"][1] + item["bbox"][3]) / 2, item["bbox"][0]))

    lines: list[dict[str, Any]] = []
    for word in prepared:
        center_y = (word["bbox"][1] + word["bbox"][3]) / 2
        placed = False
        for line in lines:
            if abs(center_y - float(line["center_y"])) <= line_tolerance:
                line["words"].append(word)
                line["center_y"] = (
                    sum((entry["bbox"][1] + entry["bbox"][3]) / 2 for entry in line["words"]) / len(line["words"])
                )
                placed = True
                break
        if not placed:
            lines.append({"center_y": center_y, "words": [word]})

    normalized_lines: list[dict[str, Any]] = []
    for line in lines:
        line_words = sorted(line["words"], key=lambda entry: entry["bbox"][0])
        segments: list[list[dict[str, Any]]] = [[line_words[0]]]
        for word in line_words[1:]:
            previous = segments[-1][-1]
            gap = float(word["bbox"][0]) - float(previous["bbox"][2])
            gap_limit = max(median_height * (1.8 if layout_hint in {"diagram", "mixed"} else 2.1), image_width * (0.035 if layout_hint in {"diagram", "mixed"} else 0.045))
            if gap > gap_limit:
                segments.append([word])
            else:
                segments[-1].append(word)
        text = "  ".join(" ".join(entry["text"] for entry in segment).strip() for segment in segments if segment).strip()
        if not text:
            continue
        x0 = min(entry["bbox"][0] for entry in line_words)
        x1 = max(entry["bbox"][2] for entry in line_words)
        y0 = min(entry["bbox"][1] for entry in line_words)
        y1 = max(entry["bbox"][3] for entry in line_words)
        normalized_lines.append(
            {
                "text": text,
                "x0": x0,
                "x1": x1,
                "y0": y0,
                "y1": y1,
                "center_x": (x0 + x1) / 2,
                "center_y": (y0 + y1) / 2,
                "width": x1 - x0,
            }
        )
    if not normalized_lines:
        return ""

    column_lines = [line for line in normalized_lines if line["width"] <= image_width * 0.7]
    column_centers = _cluster_reconstructed_positions(
        [float(line["center_x"]) for line in column_lines],
        tolerance=image_width * 0.18,
    )
    if len(column_centers) > 3:
        column_centers = column_centers[:3]

    if layout_hint not in {"diagram", "mixed"}:
        column_centers = []

    for line in normalized_lines:
        if not column_centers:
            line["column"] = 0
        else:
            line["column"] = min(range(len(column_centers)), key=lambda idx: abs(column_centers[idx] - float(line["center_x"])))

    ordered = sorted(normalized_lines, key=lambda line: (int(line["column"]), float(line["y0"]), float(line["x0"])))
    parts: list[str] = []
    previous_by_column: dict[int, float] = {}
    last_column: int | None = None
    for line in ordered:
        column = int(line["column"])
        if last_column is not None and column != last_column:
            parts.append("")
        gap = float(line["y0"]) - float(previous_by_column.get(column, line["y0"]))
        if column in previous_by_column and gap > max(median_height * 1.9, image_height * 0.028):
            parts.append("")
        parts.append(str(line["text"]))
        previous_by_column[column] = float(line["y1"])
        last_column = column
    return _clean_pdf_text("\n".join(parts))


def _extract_label_retry_words_from_image(
    image: Any,
    diagram: dict[str, Any],
    profile: dict[str, Any],
    *,
    page_width: float,
    page_height: float,
    layout_hint: str,
) -> tuple[list[dict[str, Any]], list[str], dict[str, Any]]:
    if layout_hint not in {"diagram", "mixed"}:
        return [], [], {}
    boxes = [box for box in diagram.get("boxes", []) if isinstance(box, dict) and not str(box.get("label", "")).strip()]
    if not boxes:
        return [], [], {}

    try:
        from PIL import Image
        import pytesseract
    except ImportError:
        return [], [], {}

    image_width, image_height = image.size
    if image_width <= 0 or image_height <= 0 or page_width <= 0 or page_height <= 0:
        return [], [], {}

    scale_x = image_width / page_width
    scale_y = image_height / page_height
    retry_words: list[dict[str, Any]] = []
    retry_texts: list[str] = []
    attempted = 0

    candidate_boxes = sorted(
        boxes,
        key=lambda box: (
            float(box.get("bbox", [0.0, 0.0, 0.0, 0.0])[2]) - float(box.get("bbox", [0.0, 0.0, 0.0, 0.0])[0]),
            float(box.get("bbox", [0.0, 0.0, 0.0, 0.0])[3]) - float(box.get("bbox", [0.0, 0.0, 0.0, 0.0])[1]),
        ),
    )[:8]
    for box in candidate_boxes:
        bbox = box.get("bbox", [])
        if not isinstance(bbox, list) or len(bbox) < 4:
            continue
        x0, y0, x1, y1 = [float(value) for value in bbox[:4]]
        if x1 <= x0 or y1 <= y0:
            continue
        attempted += 1
        pad_x = max(10.0, (x1 - x0) * 0.18)
        pad_y = max(8.0, (y1 - y0) * 0.22)
        crop = (
            max(0, int((x0 - pad_x) * scale_x)),
            max(0, int((y0 - pad_y) * scale_y)),
            min(image_width, int((x1 + pad_x) * scale_x)),
            min(image_height, int((y1 + pad_y) * scale_y)),
        )
        if crop[2] - crop[0] < 10 or crop[3] - crop[1] < 10:
            continue
        crop_image = image.crop(crop)
        crop_profile = {
            **profile,
            "variant": "label-retry",
            "threshold": max(int(profile.get("threshold", 0) or 0) - 36, 92),
            "word_confidence_floor": min(int(profile.get("word_confidence_floor", 40) or 40), 16),
        }
        processed = _preprocess_ocr_image(crop_image, crop_profile)
        crop_data = pytesseract.image_to_data(
            processed,
            lang="kor+eng",
            config="--psm 6",
            output_type=pytesseract.Output.DICT,
        )
        crop_words = _extract_ocr_words(
            crop_data,
            min_confidence=float(crop_profile.get("word_confidence_floor", 16) or 16),
            layout_hint=layout_hint,
            label_focus=True,
        )
        if not crop_words:
            continue
        offset_x, offset_y = crop[0], crop[1]
        for word in crop_words:
            bbox_word = word.get("bbox", [])
            if not isinstance(bbox_word, list) or len(bbox_word) != 4:
                continue
            word["bbox"] = [
                int(bbox_word[0] + offset_x),
                int(bbox_word[1] + offset_y),
                int(bbox_word[2] + offset_x),
                int(bbox_word[3] + offset_y),
            ]
        retry_words.extend(crop_words)
        retry_texts.extend(word["text"] for word in crop_words if word.get("text"))

    if not retry_words:
        return [], [], {"attempted_boxes": attempted, "recovered_boxes": 0}
    return retry_words, retry_texts, {"attempted_boxes": attempted, "recovered_boxes": len(retry_words)}


def _merge_ocr_word_lists(
    base_words: list[dict[str, Any]],
    extra_words: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    merged = [dict(word) for word in base_words if isinstance(word, dict)]
    for word in extra_words:
        if not isinstance(word, dict):
            continue
        text = _normalize_ocr_word_text(str(word.get("text", "")))
        bbox = word.get("bbox", [])
        if not text or not isinstance(bbox, list) or len(bbox) != 4:
            continue
        candidate = tuple(float(value) for value in bbox[:4])
        duplicate = False
        for existing in merged:
            existing_text = _normalize_ocr_word_text(str(existing.get("text", "")))
            existing_bbox = existing.get("bbox", [])
            if not existing_text or not isinstance(existing_bbox, list) or len(existing_bbox) != 4:
                continue
            if existing_text != text:
                continue
            existing_rect = tuple(float(value) for value in existing_bbox[:4])
            if _bbox_overlaps(existing_rect, candidate):
                duplicate = True
                break
        if duplicate:
            continue
        merged.append(
            {
                "text": text,
                "confidence": float(word.get("confidence", 0.0) or 0.0),
                "bbox": [int(candidate[0]), int(candidate[1]), int(candidate[2]), int(candidate[3])],
            }
        )
    return merged


def _preprocess_ocr_image(image: Any, profile: dict[str, Any]) -> Any:
    from PIL import ImageOps

    processed = image.convert("L") if profile.get("grayscale", True) else image.convert("RGB")
    processed = ImageOps.autocontrast(processed)
    threshold = int(profile.get("threshold", 0) or 0)
    if threshold > 0:
        processed = processed.point(lambda value: 255 if value >= threshold else 0)
    return processed.convert("RGB")


def _extract_ocr_from_raster(
    task: tuple[int, bytes, bool, int, str, dict[str, Any], tuple[float, float]],
) -> tuple[int, dict[str, str]]:
    page_index, image_bytes, force, _dpi, profile_key, diagram, page_size = task
    if not shutil.which("tesseract"):
        return page_index, {"text": "", "overlay": "", "confidence_summary": {}}
    try:
        from PIL import Image
        import pytesseract
    except ImportError:
        return page_index, {"text": "", "overlay": "", "confidence_summary": {}}

    try:
        image = Image.open(io.BytesIO(image_bytes)).convert("RGB")
        width, height = image.size
        layout_hint = profile_key.split("-", 1)[1] if "-" in profile_key else "text"
        profile = _build_ocr_profile(layout_hint, force=force)
        best_payload: dict[str, Any] | None = None
        trials: list[dict[str, Any]] = []
        variants = _build_ocr_profile_variants(profile)
        pending_variants = list(variants)
        stop_reason: str | None = None
        best_summary: dict[str, Any] | None = None
        tried_variants: list[str] = []
        while pending_variants:
            variant = pending_variants.pop(0)
            tried_variants.append(str(variant.get("variant", "")))
            processed = _preprocess_ocr_image(image, variant)
            for psm in profile["psm_candidates"]:
                config = f"--psm {psm}"
                text = _clean_pdf_text(pytesseract.image_to_string(processed, lang="kor+eng", config=config))
                data = pytesseract.image_to_data(
                    processed,
                    lang="kor+eng",
                    config=config,
                    output_type=pytesseract.Output.DICT,
                )
                summary = _summarize_ocr_confidence(data)
                score = float(summary.get("avg", 0.0)) + min(float(summary.get("words", 0)) / 40.0, 3.0)
                trials.append(
                    {
                        "psm": psm,
                        "variant": variant["variant"],
                        "threshold": variant["threshold"],
                        "avg": summary.get("avg", 0.0),
                        "words": summary.get("words", 0),
                        "low_confidence_ratio": summary.get("low_confidence_ratio", 0.0),
                        "score": round(score, 3),
                    }
                )
                if best_payload is None or score > float(best_payload["score"]):
                    words = _extract_ocr_words(
                        data,
                        min_confidence=float(profile.get("word_confidence_floor", 40.0) or 40.0),
                        layout_hint=layout_hint,
                    )
                    structured_text = _build_structured_ocr_text(
                        words,
                        image_width=width,
                        image_height=height,
                        layout_hint=layout_hint,
                    )
                    best_summary = summary
                    best_payload = {
                        "score": score,
                        "text": structured_text or text,
                        "overlay": _render_ocr_overlay(data, width, height),
                        "confidence_summary": summary,
                        "profile": profile["key"],
                        "words": words,
                        "image_width": width,
                        "image_height": height,
                        "strategy": {
                            "psm": psm,
                            "variant": variant["variant"],
                            "threshold": variant["threshold"],
                            "candidates": profile["psm_candidates"],
                            "selected_score": round(score, 3),
                            "trials": trials,
                        },
                    }
                stop_reason = _should_stop_ocr_trials(summary, score)
                if stop_reason:
                    break
            if stop_reason:
                break
            if not pending_variants:
                pending_variants.extend(_should_expand_ocr_variants(profile, tried_variants, best_summary))
        if stop_reason is None and best_summary is not None and len(trials) > len(profile["psm_candidates"]):
            stop_reason = "exhausted-candidates"
        if best_payload is None:
            return page_index, {"text": "", "overlay": "", "confidence_summary": {}, "profile": profile["key"], "strategy": {}}
        label_retry_info: dict[str, Any] = {}
        if (
            isinstance(diagram, dict)
            and layout_hint in {"diagram", "mixed"}
            and page_size[0] > 0
            and page_size[1] > 0
        ):
            retry_words, retry_texts, label_retry_info = _extract_label_retry_words_from_image(
                image,
                diagram,
                profile,
                page_width=float(page_size[0]),
                page_height=float(page_size[1]),
                layout_hint=layout_hint,
            )
            if retry_words:
                merged_words = _merge_ocr_word_lists(best_payload.get("words", []), retry_words)
                merged_text = _build_structured_ocr_text(
                    merged_words,
                    image_width=width,
                    image_height=height,
                    layout_hint=layout_hint,
                )
                retry_text = _clean_pdf_text("\n".join(dict.fromkeys(text for text in retry_texts if text)))
                if merged_words:
                    best_payload["words"] = merged_words
                if merged_text:
                    if not best_payload.get("text"):
                        best_payload["text"] = merged_text
                    elif len(merged_text) > len(str(best_payload.get("text", ""))) * 0.72:
                        best_payload["text"] = _clean_pdf_text(f"{best_payload['text']}\n{merged_text}")
                elif retry_text:
                    best_payload["text"] = _clean_pdf_text(
                        f"{best_payload.get('text', '')}\n{retry_text}" if best_payload.get("text") else retry_text
                    )
                label_retry_info["retry_words"] = len(retry_words)
                label_retry_info["merged_words"] = len(merged_words)
                if retry_text:
                    label_retry_info["recovered_text"] = retry_text
                if isinstance(best_payload.get("strategy"), dict):
                    best_payload["strategy"]["label_retry"] = label_retry_info
        if isinstance(best_payload.get("strategy"), dict):
            best_payload["strategy"]["trials"] = trials
            best_payload["strategy"]["stop_reason"] = stop_reason or ""
        best_payload.pop("score", None)
        return page_index, best_payload
    except Exception:
        return page_index, {"text": "", "overlay": "", "confidence_summary": {}}


def _render_pdf_page_faithful(
    page: Any,
    page_dict: dict[str, Any],
    ocr_overlay: str = "",
    *,
    source_path: Path | None = None,
    document: Any | None = None,
    page_index: int | None = None,
    cache_dir: Path | None = None,
    cache_report: dict[str, dict[str, Any]] | None = None,
    tables: list[dict[str, Any]] | None = None,
    ocr_status: str = "native-text",
    debug_overlays: bool = False,
) -> str:
    preview = _get_pdf_page_preview(page, dpi=144, source_path=source_path, document=document, page_index=page_index, cache_dir=cache_dir, cache_report=cache_report)
    if not preview:
        return ""

    width = preview["width"]
    height = preview["height"]
    text_layer = _render_pdf_text_overlay(page.rect.width, page.rect.height, page_dict)
    debug_layer = ""
    if debug_overlays:
        debug_layer = _render_pdf_debug_overlay(
            page.rect.width,
            page.rect.height,
            tables or [],
            ocr_status,
        )
    layers = "".join(part for part in [text_layer, ocr_overlay, debug_layer] if part)
    return (
        f"<div class=\"pdf-page-frame\" style=\"max-width:{width}px;\">"
        f"<img class=\"pdf-page-bg\" alt=\"PDF page preview\" src=\"data:image/png;base64,{preview['data']}\" "
        f"width=\"{width}\" height=\"{height}\" />"
        f"{layers}</div>"
    )


def _render_pdf_page_preview(
    page: Any,
    source_path: Path | None = None,
    document: Any | None = None,
    page_index: int | None = None,
    cache_dir: Path | None = None,
    cache_report: dict[str, dict[str, Any]] | None = None,
) -> str:
    preview = _get_pdf_page_preview(
        page,
        dpi=144,
        source_path=source_path,
        document=document,
        page_index=page_index,
        cache_dir=cache_dir,
        cache_report=cache_report,
    )
    if not preview:
        return ""
    return (
        "<figure class=\"pdf-preview\">"
        f"<img alt=\"PDF page preview\" src=\"data:image/png;base64,{preview['data']}\" "
        f"width=\"{preview['width']}\" height=\"{preview['height']}\" />"
        "</figure>"
    )


def _cluster_reconstructed_positions(values: list[float], *, tolerance: float) -> list[float]:
    if not values:
        return []
    ordered = sorted(values)
    clusters: list[list[float]] = [[ordered[0]]]
    for value in ordered[1:]:
        if abs(value - clusters[-1][-1]) <= tolerance:
            clusters[-1].append(value)
        else:
            clusters.append([value])
    return [sum(cluster) / len(cluster) for cluster in clusters]


def _annotate_reconstructed_item(page_width: float, page_height: float, item: dict[str, Any]) -> dict[str, Any]:
    x0, y0, x1, y1 = item["bbox"]
    width = max(0.0, x1 - x0)
    height = max(0.0, y1 - y0)
    center_x = (x0 + x1) / 2
    center_y = (y0 + y1) / 2
    item["width_ratio"] = width / page_width
    item["height_ratio"] = height / page_height
    item["center_x"] = center_x / page_width
    item["center_y"] = center_y / page_height
    item["center_y_abs"] = center_y
    html = str(item.get("html", ""))
    item["is_heading"] = "<h3>" in html or "<h4>" in html
    item["is_list"] = "<ul>" in html or "<ol>" in html
    item["is_table"] = item.get("kind") == "table"
    return item


def _should_skip_reconstructed_item(item: dict[str, Any], *, page_height: float) -> bool:
    if item.get("is_table") or item.get("is_heading"):
        return False
    html = str(item.get("html", ""))
    text = re.sub(r"<[^>]+>", " ", html)
    text = re.sub(r"\s+", " ", text).strip()
    if not text:
        return True
    if (
        item.get("kind") == "text"
        and float(item["bbox"][1]) >= page_height * 0.82
        and float(item.get("width_ratio", 0.0)) >= 0.72
        and float(item.get("height_ratio", 1.0)) <= 0.045
        and len(text) >= 28
    ):
        return True
    return False


def _reconstructed_items_share_row(item: dict[str, Any], row: list[dict[str, Any]]) -> bool:
    row_top = min(float(entry["bbox"][1]) for entry in row)
    row_bottom = max(float(entry["bbox"][3]) for entry in row)
    row_height = max(row_bottom - row_top, 1.0)
    row_center = sum(float(entry.get("center_y_abs", (entry["bbox"][1] + entry["bbox"][3]) / 2)) for entry in row) / len(row)
    item_top = float(item["bbox"][1])
    item_bottom = float(item["bbox"][3])
    item_center = float(item.get("center_y_abs", (item_top + item_bottom) / 2))
    overlap = min(item_bottom, row_bottom) - max(item_top, row_top)
    if overlap >= min(item_bottom - item_top, row_height) * 0.2:
        return True
    return abs(item_center - row_center) <= max(8.0, row_height * 0.42)


def _resolve_reconstructed_grid_span(
    item: dict[str, Any],
    column_centers: list[float],
    *,
    page_width: float,
) -> tuple[int, int]:
    x0, _y0, x1, _y1 = item["bbox"]
    if not column_centers:
        start_col = max(1, min(12, int((x0 / page_width) * 12) + 1))
        end_col = max(start_col + 1, min(13, int(((x1 / page_width) * 12) + 1.999)))
        return start_col, end_col

    normalized_centers = [center / page_width for center in column_centers]
    center_x = float(item.get("center_x", 0.5))
    nearest_index = min(range(len(normalized_centers)), key=lambda idx: abs(normalized_centers[idx] - center_x))
    span = 12
    if len(column_centers) == 2:
        span = 6
    elif len(column_centers) >= 3:
        span = 4
    start_col = 1 + nearest_index * span
    end_col = min(13, start_col + span)

    width_ratio = float(item.get("width_ratio", 1.0))
    if width_ratio >= 0.82 or item.get("is_table"):
        return 1, 13
    if width_ratio >= 0.58 and len(column_centers) >= 2:
        return 1, 13
    if width_ratio <= 0.22 and len(column_centers) == 1:
        start_col = max(2, min(9, int(center_x * 12)))
        end_col = min(13, start_col + 3)
    return start_col, end_col


def _build_reconstructed_ocr_items(
    page_width: float,
    page_height: float,
    ocr_payload: dict[str, Any],
) -> list[dict[str, Any]]:
    words = ocr_payload.get("words", [])
    image_width = max(int(ocr_payload.get("image_width", 0) or 0), 1)
    image_height = max(int(ocr_payload.get("image_height", 0) or 0), 1)
    if not isinstance(words, list) or not words:
        return []

    prepared: list[dict[str, Any]] = []
    heights: list[float] = []
    for word in words:
        if not isinstance(word, dict):
            continue
        text = str(word.get("text", "")).strip()
        bbox = word.get("bbox", [])
        if not text or not isinstance(bbox, list) or len(bbox) != 4:
            continue
        x0, y0, x1, y1 = [float(value) for value in bbox]
        if x1 <= x0 or y1 <= y0:
            continue
        heights.append(y1 - y0)
        prepared.append({"text": text, "bbox": [x0, y0, x1, y1]})
    if not prepared:
        return []

    median_height = sorted(heights)[len(heights) // 2] if heights else 12.0
    line_tolerance = max(10.0, median_height * 0.85)
    prepared.sort(key=lambda item: ((item["bbox"][1] + item["bbox"][3]) / 2, item["bbox"][0]))

    lines: list[list[dict[str, Any]]] = []
    for word in prepared:
        center_y = (word["bbox"][1] + word["bbox"][3]) / 2
        placed = False
        for line in lines:
            line_center = sum((entry["bbox"][1] + entry["bbox"][3]) / 2 for entry in line) / len(line)
            if abs(center_y - line_center) <= line_tolerance:
                line.append(word)
                placed = True
                break
        if not placed:
            lines.append([word])

    items: list[dict[str, Any]] = []
    for line in lines:
        line.sort(key=lambda entry: entry["bbox"][0])
        segments: list[list[dict[str, Any]]] = [[line[0]]]
        for word in line[1:]:
            prev = segments[-1][-1]
            gap = float(word["bbox"][0]) - float(prev["bbox"][2])
            if gap > max(median_height * 2.2, image_width * 0.05):
                segments.append([word])
            else:
                segments[-1].append(word)
        for segment in segments:
            text = " ".join(entry["text"] for entry in segment).strip()
            if not text:
                continue
            x0 = min(entry["bbox"][0] for entry in segment) / image_width * page_width
            y0 = min(entry["bbox"][1] for entry in segment) / image_height * page_height
            x1 = max(entry["bbox"][2] for entry in segment) / image_width * page_width
            y1 = max(entry["bbox"][3] for entry in segment) / image_height * page_height
            bbox = (x0, y0, x1, y1)
            is_title = y0 <= page_height * 0.14 and len(text) <= 56
            html = f"<h4>{escape(text)}</h4>" if is_title else f"<p>{escape(text)}</p>"
            items.append(
                _annotate_reconstructed_item(page_width, page_height, {
                    "bbox": bbox,
                    "kind": "text",
                    "html": html,
                    "height": max(0.0, y1 - y0),
                    "width": max(0.0, x1 - x0),
                })
            )
    return items


def _project_ocr_words_to_page(
    ocr_payload: dict[str, Any],
    *,
    page_width: float,
    page_height: float,
) -> list[dict[str, Any]]:
    words = ocr_payload.get("words", [])
    image_width = max(int(ocr_payload.get("image_width", 0) or 0), 1)
    image_height = max(int(ocr_payload.get("image_height", 0) or 0), 1)
    projected: list[dict[str, Any]] = []
    for word in words if isinstance(words, list) else []:
        if not isinstance(word, dict):
            continue
        text = str(word.get("text", "")).strip()
        bbox = word.get("bbox", [])
        if not text or not isinstance(bbox, list) or len(bbox) != 4:
            continue
        x0, y0, x1, y1 = [float(value) for value in bbox]
        if x1 <= x0 or y1 <= y0:
            continue
        projected.append(
            {
                "text": text,
                "confidence": float(word.get("confidence", 0.0) or 0.0),
                "bbox": [
                    x0 / image_width * page_width,
                    y0 / image_height * page_height,
                    x1 / image_width * page_width,
                    y1 / image_height * page_height,
                ],
            }
        )
    return projected


def _build_reconstructed_items_from_projected_words(
    projected_words: list[dict[str, Any]],
    *,
    page_width: float,
    page_height: float,
) -> list[dict[str, Any]]:
    if not projected_words:
        return []

    heights = [float(word["bbox"][3]) - float(word["bbox"][1]) for word in projected_words]
    median_height = sorted(heights)[len(heights) // 2] if heights else 12.0
    line_tolerance = max(10.0, median_height * 0.85)
    prepared = sorted(projected_words, key=lambda item: ((item["bbox"][1] + item["bbox"][3]) / 2, item["bbox"][0]))

    lines: list[list[dict[str, Any]]] = []
    for word in prepared:
        center_y = (word["bbox"][1] + word["bbox"][3]) / 2
        placed = False
        for line in lines:
            line_center = sum((entry["bbox"][1] + entry["bbox"][3]) / 2 for entry in line) / len(line)
            if abs(center_y - line_center) <= line_tolerance:
                line.append(word)
                placed = True
                break
        if not placed:
            lines.append([word])

    items: list[dict[str, Any]] = []
    for line in lines:
        line.sort(key=lambda entry: entry["bbox"][0])
        segments: list[list[dict[str, Any]]] = [[line[0]]]
        for word in line[1:]:
            previous = segments[-1][-1]
            gap = float(word["bbox"][0]) - float(previous["bbox"][2])
            if gap > max(median_height * 2.2, page_width * 0.05):
                segments.append([word])
            else:
                segments[-1].append(word)
        for segment in segments:
            text = " ".join(entry["text"] for entry in segment).strip()
            if not text:
                continue
            x0 = min(entry["bbox"][0] for entry in segment)
            y0 = min(entry["bbox"][1] for entry in segment)
            x1 = max(entry["bbox"][2] for entry in segment)
            y1 = max(entry["bbox"][3] for entry in segment)
            bbox = (x0, y0, x1, y1)
            is_title = y0 <= page_height * 0.14 and len(text) <= 56
            html = f"<h4>{escape(text)}</h4>" if is_title else f"<p>{escape(text)}</p>"
            items.append(
                _annotate_reconstructed_item(page_width, page_height, {
                    "bbox": bbox,
                    "kind": "text",
                    "html": html,
                    "height": max(0.0, y1 - y0),
                    "width": max(0.0, x1 - x0),
                })
            )
    return items


def _build_reconstructed_section_rows(
    projected_words: list[dict[str, Any]],
    *,
    page_width: float,
    page_height: float,
) -> list[list[str]]:
    if not projected_words:
        return []

    heights = [float(word["bbox"][3]) - float(word["bbox"][1]) for word in projected_words]
    median_height = sorted(heights)[len(heights) // 2] if heights else 12.0
    line_tolerance = max(10.0, median_height * 0.85)
    prepared = sorted(projected_words, key=lambda item: ((item["bbox"][1] + item["bbox"][3]) / 2, item["bbox"][0]))

    lines: list[list[dict[str, Any]]] = []
    for word in prepared:
        center_y = (word["bbox"][1] + word["bbox"][3]) / 2
        placed = False
        for line in lines:
            line_center = sum((entry["bbox"][1] + entry["bbox"][3]) / 2 for entry in line) / len(line)
            if abs(center_y - line_center) <= line_tolerance:
                line.append(word)
                placed = True
                break
        if not placed:
            lines.append([word])

    grouped_lines: list[list[tuple[float, float, list[str]]]] = []
    column_signatures: list[float] = []
    for line in lines:
        line = sorted(line, key=lambda entry: entry["bbox"][0])
        groups: list[list[dict[str, Any]]] = [[line[0]]]
        for word in line[1:]:
            previous = groups[-1][-1]
            gap = float(word["bbox"][0]) - float(previous["bbox"][2])
            if gap > max(median_height * 2.0, page_width * 0.04):
                groups.append([word])
            else:
                groups[-1].append(word)
        collapsed: list[tuple[float, float, list[str]]] = []
        for group in groups:
            x0 = min(entry["bbox"][0] for entry in group)
            x1 = max(entry["bbox"][2] for entry in group)
            collapsed.append((x0, x1, [entry["text"] for entry in group]))
        grouped_lines.append(collapsed)
        if not column_signatures and 2 <= len(collapsed) <= 8:
            column_signatures = [(x0 + x1) / 2 for x0, x1, _texts in collapsed]

    if not column_signatures:
        centers = [((x0 + x1) / 2) for line in grouped_lines for x0, x1, _texts in line if (x1 - x0) <= page_width * 0.7]
        column_signatures = _cluster_reconstructed_positions(centers, tolerance=page_width * 0.08)
    if len(column_signatures) < 2:
        return []
    if len(column_signatures) > 8:
        column_signatures = column_signatures[:8]

    rows: list[list[str]] = []
    for groups in grouped_lines:
        aligned = _align_row_to_columns(groups, column_signatures)
        if aligned:
            rows.append(aligned)
    widths = {len(row) for row in rows if row}
    if len(rows) < 2 or len(widths) != 1:
        return []
    return rows


def _extract_page_grid_segments(page: Any) -> tuple[list[tuple[float, float, float]], list[tuple[float, float, float]]]:
    horizontal: list[tuple[float, float, float]] = []
    vertical: list[tuple[float, float, float]] = []
    try:
        drawings = page.get_drawings()
    except Exception:
        return horizontal, vertical
    for drawing in drawings:
        for item in drawing.get("items", []):
            if not item or item[0] != "l" or len(item) < 3:
                continue
            p1, p2 = item[1], item[2]
            x1, y1 = float(getattr(p1, "x", 0.0)), float(getattr(p1, "y", 0.0))
            x2, y2 = float(getattr(p2, "x", 0.0)), float(getattr(p2, "y", 0.0))
            dx, dy = abs(x1 - x2), abs(y1 - y2)
            if dy <= 2.5 and dx >= 18.0:
                horizontal.append((round(min(x1, x2), 2), round(max(x1, x2), 2), round((y1 + y2) / 2, 2)))
            elif dx <= 2.5 and dy >= 18.0:
                vertical.append((round((x1 + x2) / 2, 2), round(min(y1, y2), 2), round(max(y1, y2), 2)))
    return horizontal, vertical


def _normalize_form_section_label(label: str) -> str:
    normalized = re.sub(r"\s+", "", label or "")
    if not normalized:
        return ""
    if "자격" in normalized or "면허" in normalized:
        return "자격면허"
    if "주거" in normalized or "재산" in normalized or normalized in {"주거", "재산"}:
        return "주거/재산"
    if "가족" in normalized or ("가" in normalized and "항" in normalized):
        return "가족사항"
    if "경력" in normalized or ("사" in normalized and "항" in normalized):
        return "경력사항"
    if "성적" in normalized or normalized == "점":
        return "성적"
    if "학" in normalized:
        return "학력"
    if "병" in normalized and "역" in normalized:
        return "병역"
    if "신" in normalized and "체" in normalized:
        return "신체"
    return re.sub(r"\s+", " ", label).strip()


def _build_reconstructed_section_grid_rows(
    projected_words: list[dict[str, Any]],
    horizontal_segments: list[tuple[float, float, float]],
    vertical_segments: list[tuple[float, float, float]],
    section_bbox: tuple[float, float, float, float],
) -> list[list[str]]:
    x0, y0, x1, y1 = section_bbox
    section_width = max(x1 - x0, 1.0)
    section_height = max(y1 - y0, 1.0)
    h_positions = []
    for sx0, sx1, sy in horizontal_segments:
        overlap = min(x1, sx1) - max(x0, sx0)
        if sy < y0 - 2.0 or sy > y1 + 2.0:
            continue
        if overlap >= section_width * 0.45:
            h_positions.append(sy)
    v_positions = []
    for sx, sy0, sy1 in vertical_segments:
        overlap = min(y1, sy1) - max(y0, sy0)
        if sx < x0 - 2.0 or sx > x1 + 2.0:
            continue
        if overlap >= section_height * 0.2:
            v_positions.append(sx)

    h_positions = _cluster_reconstructed_positions(h_positions + [y0, y1], tolerance=3.0)
    v_positions = _cluster_reconstructed_positions(v_positions + [x0, x1], tolerance=3.0)
    h_positions = sorted(value for value in h_positions if y0 - 1.0 <= value <= y1 + 1.0)
    v_positions = sorted(value for value in v_positions if x0 - 1.0 <= value <= x1 + 1.0)
    if len(h_positions) < 3 or len(v_positions) < 3:
        return []

    rows: list[list[str]] = []
    for top, bottom in zip(h_positions, h_positions[1:]):
        if bottom - top < 9.0:
            continue
        current_row: list[str] = []
        row_words = [
            word for word in projected_words
            if ((float(word["bbox"][1]) + float(word["bbox"][3])) / 2) >= top
            and ((float(word["bbox"][1]) + float(word["bbox"][3])) / 2) < bottom
        ]
        for left, right in zip(v_positions, v_positions[1:]):
            if right - left < 12.0:
                continue
            cell_words = []
            for word in row_words:
                center_x = (float(word["bbox"][0]) + float(word["bbox"][2])) / 2
                if center_x >= left and center_x < right:
                    cell_words.append(word)
            cell_words.sort(key=lambda word: (word["bbox"][1], word["bbox"][0]))
            text = " ".join(str(word["text"]) for word in cell_words).strip()
            current_row.append(text)
        if any(cell.strip() for cell in current_row):
            rows.append(current_row)
    widths = {len(row) for row in rows if row}
    if len(rows) < 2 or len(widths) != 1:
        return []
    return rows


def _extract_reconstructed_page_title(
    ocr_payload: dict[str, Any] | None,
    *,
    page_width: float,
    page_height: float,
    form_template: bool,
) -> str:
    if not ocr_payload:
        return "이력서" if form_template else ""
    words = ocr_payload.get("words", [])
    if not isinstance(words, list) or not words:
        return "이력서" if form_template else ""

    top_words = []
    for word in words:
        if not isinstance(word, dict):
            continue
        bbox = word.get("bbox", [])
        text = str(word.get("text", "")).strip()
        if not text or not isinstance(bbox, list) or len(bbox) != 4:
            continue
        x0, y0, x1, y1 = [float(value) for value in bbox]
        if y1 > page_height * 0.18 or x1 < page_width * 0.2 or x0 > page_width * 0.8:
            continue
        top_words.append((y0, x0, text, float(word.get("confidence", 0.0) or 0.0)))
    top_words.sort(key=lambda item: (item[0], item[1]))
    if not top_words:
        return "이력서" if form_template else ""

    text = "".join(item[2] for item in top_words if len(item[2]) <= 3 and not re.fullmatch(r"[A-Za-z0-9]+", item[2]))
    text = re.sub(r"\s+", "", text)
    if text and re.fullmatch(r"[가-힣]{2,6}", text):
        return text
    return "이력서" if form_template else text


def _normalize_resume_section_label(label: str, *, y0: float, page_height: float) -> str:
    cleaned = re.sub(r"\s+", "", label)
    known = {
        "학": "학력",
        "학력": "학력",
        "점": "성적",
        "성적": "성적",
        "za사항": "경력사항",
        "경력사항": "경력사항",
        "가Al항": "가족사항",
        "가족사항": "가족사항",
        "주거": "주거",
    }
    if cleaned in known:
        return known[cleaned]
    if y0 < page_height * 0.48:
        return "학력"
    if y0 < page_height * 0.60:
        return "성적"
    if y0 < page_height * 0.73:
        return "경력사항"
    if y0 < page_height * 0.89:
        return "가족사항"
    return "주거"


def _render_reconstructed_form_table(
    page: Any,
    diagram: dict[str, Any] | None,
    ocr_payload: dict[str, Any] | None,
) -> str:
    page_width = max(float(page.rect.width or 0.0), 1.0)
    page_height = max(float(page.rect.height or 0.0), 1.0)
    projected_words = _project_ocr_words_to_page(ocr_payload or {}, page_width=page_width, page_height=page_height)
    horizontal_segments, vertical_segments = _extract_page_grid_segments(page)
    if not projected_words:
        return ""

    section_boxes = []
    for box in (diagram or {}).get("boxes", []):
        if not isinstance(box, dict):
            continue
        bbox = box.get("bbox", [])
        if not isinstance(bbox, list) or len(bbox) != 4:
            continue
        x0, y0, x1, y1 = [float(value) for value in bbox]
        if x0 > page_width * 0.18 or (x1 - x0) > page_width * 0.38 or (y1 - y0) < page_height * 0.04:
            continue
        section_boxes.append((x0, y0, x1, y1, str(box.get("label", "")).strip()))
    section_boxes.sort(key=lambda item: (item[1], item[0]))
    if not section_boxes:
        return ""

    top_cutoff = section_boxes[0][1]
    top_words = [
        word
        for word in projected_words
        if float(word["bbox"][1]) <= top_cutoff - page_height * 0.02
    ]
    title = _extract_reconstructed_page_title(
        ocr_payload,
        page_width=page_width,
        page_height=page_height,
        form_template=True,
    )

    header_grid_top = None
    for _x0, _x1, sy in horizontal_segments:
        if sy < top_cutoff - page_height * 0.02:
            if header_grid_top is None or sy < header_grid_top:
                header_grid_top = sy
    if header_grid_top is None:
        header_grid_top = page_height * 0.15
    top_band_words = [
        word
        for word in projected_words
        if header_grid_top - page_height * 0.01 <= float(word["bbox"][1]) <= top_cutoff + page_height * 0.01
    ]
    top_band_rows = _build_reconstructed_section_grid_rows(
        top_band_words,
        horizontal_segments,
        vertical_segments,
        (page_width * 0.04, header_grid_top, page_width * 0.96, top_cutoff + page_height * 0.01),
    )
    if not top_band_rows:
        top_band_rows = _build_reconstructed_section_rows(
            top_band_words,
            page_width=page_width,
            page_height=page_height,
        )
    top_table = ""
    if top_band_rows:
        top_table = "<section class=\"pdf-form-overview\">" + _render_table_rows(top_band_rows, None, inferred=False)[0] + "</section>"

    section_html: list[str] = []
    for index, (x0, y0, x1, y1, raw_label) in enumerate(section_boxes):
        label = _normalize_resume_section_label(raw_label, y0=y0, page_height=page_height)
        next_top = section_boxes[index + 1][1] if index + 1 < len(section_boxes) else page_height * 0.97
        content_words = [
            word
            for word in projected_words
            if float(word["bbox"][1]) >= y0 - page_height * 0.01
            and float(word["bbox"][3]) <= next_top + page_height * 0.01
            and float(word["bbox"][0]) >= min(x1 + page_width * 0.01, page_width * 0.42)
        ]
        rows = _build_reconstructed_section_grid_rows(
            content_words,
            horizontal_segments,
            vertical_segments,
            (min(x1 + page_width * 0.01, page_width * 0.42), y0, page_width * 0.98, next_top),
        )
        if not rows:
            rows = _build_reconstructed_section_rows(
                content_words,
                page_width=page_width,
                page_height=page_height,
            )
        table_html = ""
        if rows:
            table_html = _render_table_rows(rows, None, inferred=False)[0]
        if not table_html:
            continue
        section_html.append(
            "<section class=\"pdf-form-section\">"
            f"<header><h2>{escape(label)}</h2></header>"
            f"{table_html}"
            "</section>"
        )

    if not section_html and not top_table:
        return ""

    return (
        "<article class=\"pdf-reconstructed pdf-reconstructed-form\">"
        f"<header class=\"pdf-form-title\"><h1>{escape(title)}</h1></header>"
        "<main class=\"pdf-reconstructed-page pdf-reconstructed-form\">"
        f"{top_table}"
        f"{''.join(section_html)}"
        "</main></article>"
    )


def _extract_reconstructed_form_sections(
    diagram: dict[str, Any] | None,
    ocr_payload: dict[str, Any] | None,
    *,
    page: Any | None = None,
    page_width: float,
    page_height: float,
) -> list[dict[str, Any]]:
    if not diagram or not ocr_payload:
        return []

    boxes = diagram.get("boxes", [])
    if not isinstance(boxes, list):
        return []
    section_boxes = []
    normalized_boxes: list[dict[str, Any]] = []
    for box in boxes:
        if not isinstance(box, dict):
            continue
        bbox = box.get("bbox", [])
        label = str(box.get("label", "")).strip()
        if not label or not isinstance(bbox, list) or len(bbox) != 4:
            continue
        x0, y0, x1, y1 = [float(value) for value in bbox]
        width = x1 - x0
        height = y1 - y0
        normalized_boxes.append({"label": label, "bbox": [x0, y0, x1, y1]})
        if x0 > page_width * 0.18:
            continue
        if width > page_width * 0.34:
            continue
        if height < page_height * 0.04:
            continue
        if not label:
            related_labels = [
                str(other["label"])
                for other in normalized_boxes
                if other["label"]
                and abs(float(other["bbox"][1]) - y0) <= page_height * 0.03
                and float(other["bbox"][0]) <= page_width * 0.55
                and float(other["bbox"][0]) >= page_width * 0.3
            ]
            if any("재" in candidate or "산" in candidate for candidate in related_labels):
                label = "주거/재산"
            elif y0 >= page_height * 0.82:
                label = "주거/재산"
        if not label:
            continue
        section_boxes.append({"label": re.sub(r"\s+", " ", label), "bbox": [x0, y0, x1, y1]})
    section_boxes.sort(key=lambda item: (item["bbox"][1], item["bbox"][0]))
    if len(section_boxes) < 3:
        return []

    projected_words = _project_ocr_words_to_page(ocr_payload, page_width=page_width, page_height=page_height)
    horizontal_segments: list[tuple[float, float, float]] = []
    vertical_segments: list[tuple[float, float, float]] = []
    if page is not None:
        horizontal_segments, vertical_segments = _extract_page_grid_segments(page)
    sections: list[dict[str, Any]] = []
    if section_boxes:
        top_limit = float(section_boxes[0]["bbox"][1])
        top_words = [
            word for word in projected_words
            if float(word["bbox"][1]) < top_limit - page_height * 0.01
        ]
        top_rows = []
        if page is not None and top_words:
            top_rows = _build_reconstructed_section_grid_rows(
                top_words,
                horizontal_segments,
                vertical_segments,
                (page_width * 0.05, page_height * 0.1, page_width * 0.96, top_limit - page_height * 0.01),
            )
        if top_rows:
            sections.append(
                {
                    "label": "인적사항",
                    "bbox": [page_width * 0.05, page_height * 0.1, page_width * 0.96, top_limit - page_height * 0.01],
                    "items": [],
                    "rows": top_rows,
                }
            )
    for index, box in enumerate(section_boxes):
        y0 = float(box["bbox"][1])
        y1 = float(box["bbox"][3])
        next_top = float(section_boxes[index + 1]["bbox"][1]) if index + 1 < len(section_boxes) else page_height * 0.97
        section_bottom = max(y1, next_top)
        content_left = min(page_width * 0.22, float(box["bbox"][2]) + page_width * 0.015)
        section_words = [
            word
            for word in projected_words
            if float(word["bbox"][0]) >= content_left
            and float(word["bbox"][1]) >= y0 - page_height * 0.008
            and float(word["bbox"][3]) <= section_bottom + page_height * 0.01
        ]
        items = _build_reconstructed_items_from_projected_words(
            section_words,
            page_width=page_width,
            page_height=page_height,
        )
        rows = _build_reconstructed_section_grid_rows(
            section_words,
            horizontal_segments,
            vertical_segments,
            (content_left, y0, page_width * 0.98, section_bottom),
        )
        if not rows:
            rows = _build_reconstructed_section_rows(
                section_words,
                page_width=page_width,
                page_height=page_height,
            )
        if not items:
            continue
        sections.append(
            {
                "label": _normalize_form_section_label(str(box["label"])),
                "bbox": [content_left, y0, page_width * 0.98, section_bottom],
                "items": items,
                "rows": rows,
            }
        )
    return sections


def _infer_reconstructed_page_kind(items: list[dict[str, Any]], ocr_payload: dict[str, Any] | None) -> str:
    if not items:
        return "content"
    tables = sum(1 for item in items if item.get("is_table"))
    headings = sum(1 for item in items if item.get("is_heading"))
    narrow = sum(1 for item in items if float(item.get("width_ratio", 1.0)) <= 0.34)
    if ocr_payload and not tables and len(items) >= 14 and narrow >= max(8, len(items) // 2) and headings <= max(4, len(items) // 8):
        return "form"
    return "content"


def _reconstructed_row_tag(page_kind: str, row: list[dict[str, Any]]) -> str:
    if row and all(item.get("is_heading") for item in row):
        return "header"
    if page_kind == "form":
        return "fieldset"
    return "section"


def _reconstructed_item_tag(page_kind: str, item: dict[str, Any]) -> str:
    if item.get("is_heading"):
        return "div"
    if item.get("is_table"):
        return "section"
    if page_kind == "form":
        return "article"
    return "article"


def _render_pdf_page_reconstructed(
    page: Any,
    blocks: list[dict[str, Any]],
    tables: list[dict[str, Any]],
    table_bboxes: list[tuple[float, float, float, float]],
    semantic_body: str,
    ocr_payload: dict[str, Any] | None = None,
    diagram: dict[str, Any] | None = None,
) -> str:
    page_width = max(float(page.rect.width or 0.0), 1.0)
    page_height = max(float(page.rect.height or 0.0), 1.0)
    items: list[dict[str, Any]] = []

    for block in blocks:
        if block.get("type") != 0:
            continue
        bbox = tuple(float(value) for value in block.get("bbox", (0.0, 0.0, 0.0, 0.0)))
        if any(_bbox_overlaps(bbox, table_bbox) for table_bbox in table_bboxes):
            continue
        _text, block_html, _markdown = _render_pymupdf_text_block(block)
        if not block_html:
            continue
        items.append(
            _annotate_reconstructed_item(page_width, page_height, {
                "bbox": bbox,
                "kind": "text",
                "html": block_html,
                "height": max(0.0, bbox[3] - bbox[1]),
                "width": max(0.0, bbox[2] - bbox[0]),
            })
        )

    for table in tables:
        bbox = tuple(float(value) for value in table.get("bbox", (0.0, 0.0, 0.0, 0.0)))
        if not table.get("html"):
            continue
        items.append(
            _annotate_reconstructed_item(page_width, page_height, {
                "bbox": bbox,
                "kind": "table",
                "html": table["html"],
                "height": max(0.0, bbox[3] - bbox[1]),
                "width": max(0.0, bbox[2] - bbox[0]),
            })
        )

    if not items and ocr_payload:
        items = _build_reconstructed_ocr_items(page_width, page_height, ocr_payload)
    items = [item for item in items if not _should_skip_reconstructed_item(item, page_height=page_height)]
    items.sort(key=lambda item: (item["bbox"][1], item["bbox"][0]))
    if not items:
        if semantic_body:
            return (
                "<article class=\"pdf-reconstructed pdf-reconstructed-content\">"
                "<div class=\"pdf-reconstructed-page\">"
                f"{semantic_body}</div></article>"
            )
        return ""

    page_kind = _infer_reconstructed_page_kind(items, ocr_payload)
    rows: list[list[dict[str, Any]]] = []
    for item in items:
        placed = False
        for row in rows:
            if _reconstructed_items_share_row(item, row):
                row.append(item)
                placed = True
                break
        if not placed:
            rows.append([item])

    if page_kind == "form":
        form_html = _render_reconstructed_form_table(page, diagram, ocr_payload)
        if form_html:
            return form_html

    row_html: list[str] = []
    previous_bottom = 0.0
    for row_index, row in enumerate(rows, start=1):
        row.sort(key=lambda entry: (entry["bbox"][0], entry["bbox"][1]))
        row_top = min(entry["bbox"][1] for entry in row)
        row_bottom = max(entry["bbox"][3] for entry in row)
        row_height = max(row_bottom - row_top, 1.0)
        gap = max(0.0, row_top - previous_bottom)
        row_style = ""
        if gap > 4.0:
            margin_top = min(max(gap / page_height * 280.0, 0.0), 28.0)
            if margin_top > 0.0:
                row_style = f" style=\"margin-top:{margin_top:.1f}px;\""
        row_column_centers = _cluster_reconstructed_positions(
            [float(entry["bbox"][0] + entry["bbox"][2]) / 2 for entry in row if float(entry.get("width_ratio", 1.0)) < 0.72],
            tolerance=page_width * 0.14,
        )
        if len(row_column_centers) > 3:
            row_column_centers = row_column_centers[:3]
        row_classes = ["pdf-reconstructed-row"]
        if len(row_column_centers) >= 2:
            row_classes.append(f"columns-{min(len(row_column_centers), 3)}")
        row_tag = _reconstructed_row_tag(page_kind, row)
        if page_kind == "form" and row_tag == "fieldset":
            row_classes.append("form-row")
        parts: list[str] = []
        for item in row:
            start_col, end_col = _resolve_reconstructed_grid_span(item, row_column_centers, page_width=page_width)
            classes = ["pdf-reconstructed-item", item["kind"]]
            if item["width"] <= page_width * 0.28:
                classes.append("narrow")
            if item["is_heading"] and float(item.get("center_x", 0.5)) >= 0.25 and float(item.get("center_x", 0.5)) <= 0.75:
                classes.append("centered")
            if item["width"] >= page_width * 0.82 or item["is_table"]:
                classes.append("full-width")
            if item["kind"] == "text" and item["height"] <= page_height * 0.08 and item["width"] <= page_width * 0.7 and not item["is_heading"]:
                classes.append("callout")
            if len(row) >= 2 and row_height <= page_height * 0.12 and item["width"] <= page_width * 0.32:
                classes.append("narrow")
            if page_kind == "form" and not item.get("is_heading") and not item.get("is_table"):
                classes.append("form-field")
            item_tag = _reconstructed_item_tag(page_kind, item)
            parts.append(
                f"<{item_tag} class=\"{' '.join(classes)}\" style=\"grid-column:{start_col} / {end_col};\">{item['html']}</{item_tag}>"
            )
        if row_tag == "fieldset":
            legend = f"<legend>Field Group {row_index}</legend>" if row_index <= 3 else ""
            row_html.append(f"<fieldset class=\"{' '.join(row_classes)}\"{row_style}>{legend}{''.join(parts)}</fieldset>")
        else:
            row_html.append(f"<{row_tag} class=\"{' '.join(row_classes)}\"{row_style}>{''.join(parts)}</{row_tag}>")
        previous_bottom = row_bottom

    summary = (
        "<header class=\"pdf-reconstructed-summary\">"
        f"<p>Reconstructed DOM mode: {'form-like' if page_kind == 'form' else 'content-focused'} page layout.</p>"
        "</header>"
    )
    page_tag = "form" if page_kind == "form" else "main"
    page_class = f"pdf-reconstructed-page pdf-reconstructed-{page_kind}"
    return (
        f"<article class=\"pdf-reconstructed pdf-reconstructed-{page_kind}\">"
        f"{summary}"
        f"<{page_tag} class=\"{page_class}\">{''.join(row_html)}</{page_tag}>"
        "</article>"
    )


def _get_pdf_page_preview(
    page: Any,
    dpi: int,
    *,
    source_path: Path | None = None,
    document: Any | None = None,
    page_index: int | None = None,
    cache_dir: Path | None = None,
    cache_report: dict[str, dict[str, Any]] | None = None,
) -> dict[str, Any] | None:
    if source_path and document is not None and page_index is not None:
        cached = _get_cached_page_raster(page, source_path, document, page_index, dpi, cache_dir, cache_report)
        if cached:
            return cached
    try:
        pixmap = page.get_pixmap(dpi=dpi, alpha=False)
    except Exception:
        return None
    png_bytes = pixmap.tobytes("png")
    return {
        "data": base64.b64encode(png_bytes).decode("ascii"),
        "width": pixmap.width,
        "height": pixmap.height,
    }


def _get_cached_page_raster(
    page: Any,
    source_path: Path,
    document: Any,
    page_index: int,
    dpi: int,
    cache_dir: Path | None,
    cache_report: dict[str, dict[str, Any]] | None = None,
) -> dict[str, Any] | None:
    cache_key = _build_page_cache_key("raster", source_path, page_index, dpi=dpi)
    cache_meta = _build_page_cache_metadata(source_path, document, page_index, dpi=dpi)
    cached = _load_json_cache(cache_dir, "raster", cache_key, cache_meta, cache_report)
    if isinstance(cached, dict):
        data = str(cached.get("data", ""))
        width = int(cached.get("width", 0) or 0)
        height = int(cached.get("height", 0) or 0)
        if data and width > 0 and height > 0:
            return {"data": data, "width": width, "height": height}
    cache_path = _cache_file_path(cache_dir, "raster", cache_key)
    with _cache_lock(cache_path):
        if cache_path and cache_path.exists():
            cached = _load_json_cache_from_path(cache_path, "raster", cache_meta, cache_report)
            if isinstance(cached, dict):
                data = str(cached.get("data", ""))
                width = int(cached.get("width", 0) or 0)
                height = int(cached.get("height", 0) or 0)
                if data and width > 0 and height > 0:
                    return {"data": data, "width": width, "height": height}
        try:
            pixmap = page.get_pixmap(dpi=dpi, alpha=False)
        except Exception:
            return None
        payload = {
            "data": base64.b64encode(pixmap.tobytes("png")).decode("ascii"),
            "width": pixmap.width,
            "height": pixmap.height,
        }
        _save_json_cache(cache_dir, "raster", cache_key, cache_meta, payload, cache_report)
        return payload


def _normalize_cached_table_entry(payload: Any) -> dict[str, Any]:
    if not isinstance(payload, dict):
        return {"bbox": (0.0, 0.0, 0.0, 0.0), "html": "", "text": "", "markdown": "", "source": "unknown"}
    bbox = payload.get("bbox", (0.0, 0.0, 0.0, 0.0))
    if isinstance(bbox, list):
        bbox = tuple(float(value) for value in bbox[:4])
    return {
        "bbox": bbox,
        "html": str(payload.get("html", "")),
        "text": str(payload.get("text", "")),
        "markdown": str(payload.get("markdown", "")),
        "confidence": payload.get("confidence"),
        "source": str(payload.get("source", "unknown")),
        "column_signatures": list(payload.get("column_signatures", [])),
    }


def _render_pdf_text_overlay(page_width: float, page_height: float, page_dict: dict[str, Any] | None = None) -> str:
    blocks_dict = page_dict or {}
    blocks = blocks_dict.get("blocks", [])
    if not blocks:
        return ""

    overlay_parts: list[str] = []
    for block in blocks:
        if block.get("type") != 0:
            continue
        for line in block.get("lines", []):
            for span in line.get("spans", []):
                text = span.get("text", "")
                if not text or not text.strip():
                    continue
                x0, y0, x1, y1 = [float(v) for v in span.get("bbox", [0.0, 0.0, 0.0, 0.0])]
                font_size = float(span.get("size", 10.0))
                font_flags = int(span.get("flags", 0))
                text_width = max(x1 - x0, 1.0)
                estimated_width = max(font_size * max(len(text.strip()), 1) * 0.5, 1.0)
                scale_x = max(min(text_width / estimated_width, 6.0), 0.5)
                style = (
                    f"left:{(x0 / page_width) * 100:.4f}%;"
                    f"top:{(y0 / page_height) * 100:.4f}%;"
                    f"width:{((x1 - x0) / page_width) * 100:.4f}%;"
                    f"font-size:{(font_size / page_height) * 100:.4f}%;"
                    f"transform:scaleX({scale_x:.4f});"
                )
                if font_flags & 16:
                    style += "font-weight:700;"
                if font_flags & 2:
                    style += "font-style:italic;"
                overlay_parts.append(f"<span class=\"pdf-text-span\" style=\"{style}\">{escape(text)}</span>")

    if not overlay_parts:
        return ""
    return f"<div class=\"pdf-text-layer\" aria-hidden=\"true\">{''.join(overlay_parts)}</div>"


def _render_ocr_overlay(ocr_data: dict[str, list[Any]], image_width: int, image_height: int) -> str:
    total = len(ocr_data.get("text", []))
    if not total:
        return ""

    parts: list[str] = []
    for index in range(total):
        text = (ocr_data["text"][index] or "").strip()
        confidence = ocr_data.get("conf", ["-1"] * total)[index]
        try:
            score = float(confidence)
        except (TypeError, ValueError):
            score = -1.0
        if not text or score < 20:
            continue

        left = int(ocr_data["left"][index])
        top = int(ocr_data["top"][index])
        width = int(ocr_data["width"][index])
        height = int(ocr_data["height"][index])
        if width <= 0 or height <= 0:
            continue

        font_size = max(height * 0.85, 8)
        estimated_width = max(font_size * max(len(text.strip()), 1) * 0.5, 1.0)
        scale_x = max(min(width / estimated_width, 6.0), 0.5)
        style = (
            f"left:{(left / image_width) * 100:.4f}%;"
            f"top:{(top / image_height) * 100:.4f}%;"
            f"width:{(width / image_width) * 100:.4f}%;"
            f"font-size:{(font_size / image_height) * 100:.4f}%;"
            f"transform:scaleX({scale_x:.4f});"
        )
        parts.append(f"<span class=\"pdf-text-span\" style=\"{style}\">{escape(text)}</span>")

    if not parts:
        return ""
    return f"<div class=\"pdf-text-layer\" aria-hidden=\"true\">{''.join(parts)}</div>"


def _summarize_ocr_confidence(ocr_data: dict[str, list[Any]]) -> dict[str, float | int]:
    scores: list[float] = []
    low_confidence = 0
    confident = 0
    total = len(ocr_data.get("text", []))
    for index in range(total):
        text = (ocr_data.get("text", [""] * total)[index] or "").strip()
        confidence = ocr_data.get("conf", ["-1"] * total)[index]
        try:
            score = float(confidence)
        except (TypeError, ValueError):
            continue
        if not text or score < 0:
            continue
        scores.append(score)
        if score < 50:
            low_confidence += 1
        if score >= 85:
            confident += 1
    if not scores:
        return {}
    ordered = sorted(scores)
    middle = len(ordered) // 2
    if len(ordered) % 2:
        median = ordered[middle]
    else:
        median = (ordered[middle - 1] + ordered[middle]) / 2
    return {
        "avg": round(sum(scores) / len(scores), 2),
        "median": round(median, 2),
        "min": round(min(scores), 2),
        "max": round(max(scores), 2),
        "words": len(scores),
        "confident_words": confident,
        "low_confidence_words": low_confidence,
        "low_confidence_ratio": round(low_confidence / len(scores), 3),
    }


def _render_pdf_debug_overlay(
    page_width: float,
    page_height: float,
    tables: list[dict[str, Any]],
    ocr_status: str,
) -> str:
    parts: list[str] = []
    for index, table in enumerate(tables, start=1):
        x0, y0, x1, y1 = [float(value) for value in table.get("bbox", (0.0, 0.0, 0.0, 0.0))]
        if x1 <= x0 or y1 <= y0:
            continue
        label = f"table {index}"
        if table.get("source"):
            label += f" [{table['source']}]"
        confidence = table.get("confidence")
        if isinstance(confidence, (float, int)):
            label += f" {float(confidence):.2f}"
        box_class = "pdf-debug-box"
        if table.get("source") == "inferred":
            box_class += " inferred"
        style = (
            f"left:{(x0 / page_width) * 100:.4f}%;"
            f"top:{(y0 / page_height) * 100:.4f}%;"
            f"width:{((x1 - x0) / page_width) * 100:.4f}%;"
            f"height:{((y1 - y0) / page_height) * 100:.4f}%;"
        )
        parts.append(
            f"<div class=\"{box_class}\" style=\"{style}\">"
            f"<span class=\"pdf-debug-label\">{escape(label)}</span>"
            "</div>"
        )
    if ocr_status != "native-text":
        parts.append(f"<div class=\"pdf-debug-badge\">OCR: {escape(ocr_status)}</div>")
    if not parts:
        return ""
    return f"<div class=\"pdf-debug-layer\" aria-hidden=\"true\">{''.join(parts)}</div>"


def _bbox_overlaps(a: tuple[float, float, float, float], b: tuple[float, float, float, float]) -> bool:
    ax0, ay0, ax1, ay1 = a
    bx0, by0, bx1, by1 = b
    x_overlap = max(0.0, min(ax1, bx1) - max(ax0, bx0))
    y_overlap = max(0.0, min(ay1, by1) - max(ay0, by0))
    return x_overlap > 1.0 and y_overlap > 1.0


def _clean_pdf_text(text: str) -> str:
    text = text.replace("\xa0", " ")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r" *\n *", "\n", text)
    return text.strip()


def _render_pdf_page_html(index: int, raw_text: str) -> str:
    lines = _normalize_pdf_lines(raw_text)
    blocks = _group_pdf_blocks(lines)
    rendered = [_render_pdf_block(block_type, block_lines) for block_type, block_lines in blocks]
    body = "\n".join(part for part in rendered if part)
    if not body:
        body = "<p></p>"
    return f"<section class=\"pdf-page\"><h2>Page {index}</h2>{body}</section>"


def _render_pdf_page_markdown(index: int, raw_text: str) -> str:
    lines = _normalize_pdf_lines(raw_text)
    blocks = _group_pdf_blocks(lines)
    rendered = [_render_pdf_block_markdown(block_type, block_lines) for block_type, block_lines in blocks]
    body = "\n\n".join(part for part in rendered if part).strip()
    return f"## Page {index}\n\n{body}".strip()


def _normalize_pdf_lines(raw_text: str) -> list[str]:
    normalized = []
    for raw_line in raw_text.splitlines():
        line = raw_line.replace("\xa0", " ").rstrip()
        if not line.strip():
            normalized.append("")
            continue
        if _is_pdf_noise(line.strip()):
            continue
        normalized.append(line)

    normalized = _merge_broken_pdf_lines(normalized)
    return [line for line in normalized if line.strip()]


def _is_pdf_noise(line: str) -> bool:
    if not line:
        return True
    if re.fullmatch(r"\d+", line):
        return True
    if re.fullmatch(r"Version\s+\d+(?:\.\d+)*", line):
        return True
    if re.fullmatch(r"\d+\s+Version\s+\d+(?:\.\d+)*", line):
        return True
    if "무단으로 타인에게 배포되거나" in line:
        return True
    return False


def _merge_broken_pdf_lines(lines: list[str]) -> list[str]:
    merged: list[str] = []
    index = 0
    while index < len(lines):
        current = lines[index].strip()
        if not current:
            index += 1
            continue

        while index + 1 < len(lines):
            next_line = lines[index + 1].strip()
            if not next_line or _starts_new_pdf_block(current, next_line):
                break
            if _should_merge_pdf_lines(current, next_line):
                joiner = "" if _needs_tight_join(current, next_line) else " "
                current = f"{current}{joiner}{next_line}".strip()
                index += 1
                continue
            break

        merged.append(current)
        index += 1
    return merged


def _starts_new_pdf_block(current: str, next_line: str) -> bool:
    if _looks_like_pdf_heading(next_line):
        return True
    if _looks_like_pdf_list_item(next_line):
        return True
    if _looks_like_pdf_table_row(next_line) and not _should_merge_pdf_lines(current, next_line):
        return True
    return False


def _should_merge_pdf_lines(current: str, next_line: str) -> bool:
    if not current or not next_line:
        return False
    if _looks_like_pdf_heading(current) or _looks_like_pdf_heading(next_line):
        return False
    if _looks_like_pdf_list_item(next_line):
        return False
    if _looks_like_pdf_table_row(current) or _looks_like_pdf_table_row(next_line):
        return False
    if _looks_like_pdf_list_item(current):
        return True
    if current.endswith(("(", "/", "-")) or next_line.startswith((")", ",", ".", "%", "(")):
        return True
    if current.endswith((".", "!", "?", ":", ";", ")", "]")):
        return False
    if len(current) <= 18:
        return False
    return True


def _needs_tight_join(current: str, next_line: str) -> bool:
    return current.endswith(("(", "/", "-")) or next_line.startswith((")", ",", ".", "%"))


def _group_pdf_blocks(lines: list[str]) -> list[tuple[str, list[str]]]:
    blocks: list[tuple[str, list[str]]] = []
    index = 0

    while index < len(lines):
        line = lines[index]

        if _looks_like_pdf_heading(line):
            level = _pdf_heading_level(line)
            blocks.append((f"heading-{level}", [line]))
            index += 1
            continue

        if _looks_like_pdf_list_item(line):
            items = [line]
            index += 1
            while index < len(lines) and _looks_like_pdf_list_item(lines[index]):
                items.append(lines[index])
                index += 1
            blocks.append(("list", items))
            continue

        if _looks_like_pdf_table_row(line):
            rows = [_split_pdf_columns(line)]
            index += 1
            while index < len(lines) and _looks_like_pdf_table_row(lines[index]):
                rows.append(_split_pdf_columns(lines[index]))
                index += 1
            if len(rows) >= 2:
                blocks.append(("table", ["\t".join(row) for row in rows]))
            else:
                blocks.append(("paragraph", [" ".join(rows[0])]))
            continue

        paragraph = [line]
        index += 1
        while index < len(lines):
            candidate = lines[index]
            if _looks_like_pdf_heading(candidate) or _looks_like_pdf_list_item(candidate) or _looks_like_pdf_table_row(candidate):
                break
            paragraph.append(candidate)
            index += 1
        blocks.append(("paragraph", paragraph))

    return blocks


def _render_pdf_block(block_type: str, lines: list[str]) -> str:
    if block_type.startswith("heading-"):
        level = block_type.split("-", 1)[1]
        tag = {"1": "h2", "2": "h3", "3": "h4"}.get(level, "h4")
        return f"<{tag}>{escape(lines[0])}</{tag}>"

    if block_type == "list":
        items = []
        for line in lines:
            content = re.sub(r"^[\u2022\u25aa\u25cf▪\-]+[\s]*", "", line).strip()
            content = re.sub(r"^\d+\.\s+", "", content)
            items.append(f"<li>{escape(content)}</li>")
        return "<ul>\n" + "\n".join(items) + "\n</ul>"

    if block_type == "table":
        rows = [line.split("\t") for line in lines]
        header_rows = _detect_table_header_rows(rows)
        header = _merge_table_header_band(rows[:header_rows]) if header_rows else None
        body_rows = rows[header_rows:] if header_rows else rows
        table_parts = []
        if header:
            header_html = "".join(f"<th>{escape(cell)}</th>" for cell in header)
            table_parts.append(f"<thead><tr>{header_html}</tr></thead>")
        body_html = []
        for row in body_rows:
            cells = "".join(f"<td>{escape(cell)}</td>" for cell in row)
            body_html.append(f"<tr>{cells}</tr>")
        if body_html:
            table_parts.append("<tbody>\n" + "\n".join(body_html) + "\n</tbody>")
        return "<div class=\"pdf-table-wrap\"><table>\n" + "\n".join(table_parts) + "\n</table></div>"

    content = " ".join(line.strip() for line in lines if line.strip())
    return f"<p>{escape(content)}</p>" if content else ""


def _render_pdf_block_markdown(block_type: str, lines: list[str]) -> str:
    if block_type.startswith("heading-"):
        level = block_type.split("-", 1)[1]
        prefix = {"1": "##", "2": "###", "3": "####"}.get(level, "####")
        return f"{prefix} {lines[0]}".strip()

    if block_type == "list":
        items = []
        for line in lines:
            content = re.sub(r"^[\u2022\u25aa\u25cf▪\-]+[\s]*", "", line).strip()
            content = re.sub(r"^\d+\.\s+", "", content)
            items.append(f"- {content}")
        return "\n".join(items)

    if block_type == "table":
        rows = [line.split("\t") for line in lines]
        return _render_markdown_table(rows)

    content = " ".join(line.strip() for line in lines if line.strip())
    return content


def _looks_like_pdf_heading(line: str) -> bool:
    if len(line) > 120:
        return False
    if re.match(r"^\d+(?:\.\d+){0,3}\.?\s+\S+", line):
        return True
    if re.match(r"^[A-Z][A-Za-z0-9 .&()/,-]{0,60}$", line):
        return True
    return False


def _pdf_heading_level(line: str) -> int:
    match = re.match(r"^(\d+(?:\.\d+)*)\s+", line)
    if not match:
        return 3
    depth = match.group(1).count(".")
    return min(depth + 1, 3)


def _looks_like_pdf_list_item(line: str) -> bool:
    return bool(re.match(r"^[\u2022\u25aa\u25cf▪\-]\s*", line))


def _looks_like_pdf_table_row(line: str) -> bool:
    if _looks_like_pdf_heading(line) or _looks_like_pdf_list_item(line):
        return False
    columns = _split_pdf_columns(line)
    if len(columns) < 2:
        return False
    compact = [column for column in columns if column]
    if len(compact) > 5:
        return False
    return len(compact) >= 2 and len(line) >= 12


def _split_pdf_columns(line: str) -> list[str]:
    return [part.strip() for part in re.split(r"\s{3,}", line.strip()) if part.strip()]


def _detect_table_header_rows(rows: list[list[str]] | list[str]) -> int:
    if not rows:
        return 0
    if isinstance(rows[0], str):
        candidate_rows = [rows]  # type: ignore[list-item]
    else:
        candidate_rows = rows  # type: ignore[assignment]

    header = [str(cell).strip() for cell in candidate_rows[0]]
    if len(header) < 2:
        return 0
    non_empty = [cell for cell in header if cell]
    second_row = [str(cell).strip() for cell in candidate_rows[1]] if len(candidate_rows) > 1 else []
    sparse_spanning_header = (
        len(candidate_rows) > 1
        and 1 <= len(non_empty) <= max(2, len(header) // 2)
        and any(re.search(r"[가-힣A-Za-z]", cell) for cell in non_empty)
        and sum(bool(cell) for cell in second_row) >= max(2, len(header) - 2)
    )
    if len(non_empty) < max(2, len(header) // 2) and not sparse_spanning_header:
        return 0

    header_numeric = sum(_cell_is_mostly_numeric(cell) for cell in non_empty)
    second_numeric = sum(_cell_is_mostly_numeric(cell) for cell in second_row if cell)
    header_compact = sum(len(cell) <= 24 for cell in non_empty)
    distinct_ratio = len(set(non_empty)) / max(len(non_empty), 1)

    score = 0
    if header_compact >= max(2, len(non_empty) - 1):
        score += 1
    if header_numeric == 0:
        score += 1
    if second_row and second_numeric > header_numeric:
        score += 1
    if distinct_ratio >= 0.8:
        score += 1
    if any(re.search(r"[가-힣A-Za-z]", cell) for cell in non_empty):
        score += 1
    if sparse_spanning_header:
        score += 2
    if score < 3:
        return 0
    if _looks_like_secondary_header_row(candidate_rows):
        return 2
    return 1


def _looks_like_pdf_table_header(rows: list[list[str]] | list[str]) -> bool:
    return _detect_table_header_rows(rows) >= 1


def _looks_like_secondary_header_row(rows: list[list[str]]) -> bool:
    if len(rows) < 2:
        return False
    first = [str(cell).strip() for cell in rows[0]]
    second = [str(cell).strip() for cell in rows[1]]
    second_non_empty = [cell for cell in second if cell]
    if len(second_non_empty) < max(2, len(second) // 2):
        return False
    if any(len(cell) > 32 for cell in second_non_empty):
        return False
    if sum(_cell_is_mostly_numeric(cell) for cell in second_non_empty) > 0:
        return False
    if all(not cell for cell in first[2:]) and any(cell for cell in second[2:]):
        return True
    first_non_empty = sum(bool(cell) for cell in first)
    second_fill = sum(bool(cell) for cell in second)
    if second_fill > first_non_empty and any(re.search(r"[가-힣A-Za-z]", cell) for cell in second_non_empty):
        return True
    return False


def _merge_table_header_band(rows: list[list[str]]) -> list[str]:
    if not rows:
        return []
    width = max(len(row) for row in rows)
    merged: list[str] = []
    for index in range(width):
        pieces: list[str] = []
        for row in rows:
            if index >= len(row):
                continue
            cell = str(row[index]).strip()
            if not cell or cell in pieces:
                continue
            pieces.append(cell)
        merged.append(" / ".join(pieces))
    return merged


def _cell_is_mostly_numeric(cell: str) -> bool:
    stripped = re.sub(r"[\s,./%()-]", "", cell)
    if not stripped:
        return False
    digits = sum(char.isdigit() for char in stripped)
    return digits / len(stripped) >= 0.6
